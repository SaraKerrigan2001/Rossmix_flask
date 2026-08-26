"""
Genera Excel completo del proyecto Rossmix con:
  - Portada
  - Diseño BD (10 tablas con atributos, PK/FK, relaciones)
  - Casos de Prueba (por módulo)
  - Informe de Errores del proceso
  - Software Usado
"""
import os
from datetime import date
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage

BASE = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

# ── PALETA ────────────────────────────────────────────────────────────────────
R = 'C41E3A'   # rojo Rossmix
P = 'FF1493'   # rosa vibrante
S = 'FFF0F6'   # fondo suave
W = 'FFFFFF'   # blanco
G = 'F5F5F5'   # gris muy claro
D = '1A1A1A'   # negro texto
GRIS = '6B7280'

def fill(hex6):  return PatternFill('solid', fgColor=hex6)
def font(bold=False, size=10, color=D, italic=False, name='Calibri'):
    return Font(bold=bold, size=size, color=color, italic=italic, name=name)
def align(h='left', v='center', wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)
def border(color='E5E7EB'):
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)
def border_none(): return Border()

def hdr_cell(ws, row, col, text, fill_hex=R, font_color=W, bold=True, size=10):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = fill(fill_hex)
    c.font = font(bold=bold, size=size, color=font_color)
    c.alignment = align('center','center', wrap=True)
    c.border = border('D1D5DB')
    return c

def data_cell(ws, row, col, text, fill_hex=W, color=D, bold=False, h='left', wrap=True, size=9):
    c = ws.cell(row=row, column=col, value=text)
    c.fill = fill(fill_hex)
    c.font = font(bold=bold, size=size, color=color)
    c.alignment = align(h, 'center', wrap=wrap)
    c.border = border('E5E7EB')
    return c

def merge_title(ws, row, col_start, col_end, text, fill_hex=R, font_color=W, size=11):
    ws.merge_cells(start_row=row, start_column=col_start,
                   end_row=row,   end_column=col_end)
    c = ws.cell(row=row, column=col_start, value=text)
    c.fill = fill(fill_hex)
    c.font = font(bold=True, size=size, color=font_color)
    c.alignment = align('center', 'center')
    c.border = border('C41E3A')
    ws.row_dimensions[row].height = 22
    return c

# ════════════════════════════════════════════════════════════════════════════
# DATOS — 10 TABLAS CON ATRIBUTOS
# ════════════════════════════════════════════════════════════════════════════
TABLAS = [
    {
        "nombre": "USUARIO",
        "descripcion": "Almacena todos los usuarios del sistema: clientes, administradores y especialistas. Tabla unificada con discriminador tipo_usuario.",
        "relaciones": "1:N con citas (como cliente)  ·  1:1 con empleados (especialistas)  ·  1:N con notificaciones  ·  1:N con auditoria_usuarios",
        "atributos": [
            ("PK",  "id",             "SERIAL",        "Clave primaria autoincremental"),
            ("",    "nombre",         "VARCHAR(100)",   "Nombre completo del usuario"),
            ("UK",  "email",          "VARCHAR(150)",   "Correo electrónico único del sistema"),
            ("",    "telefono",       "VARCHAR(10)",    "Celular/WhatsApp — exactamente 10 dígitos"),
            ("",    "password",       "VARCHAR(200)",   "Contraseña hasheada con Werkzeug"),
            ("",    "tipo_usuario",   "VARCHAR(20)",    "Discriminador: admin | cliente | especialista"),
            ("",    "activo",         "BOOLEAN",        "Cuenta habilitada (TRUE) o desactivada (FALSE)"),
            ("",    "fecha_registro", "TIMESTAMP",      "Fecha y hora de creación de la cuenta"),
            ("FK",  "id_empleado",    "INTEGER → empleados", "Vínculo con empleados para cuentas especialista"),
        ]
    },
    {
        "nombre": "EMPLEADOS",
        "descripcion": "Registro del personal especialista del salón. Contiene datos laborales y estado de actividad.",
        "relaciones": "1:N con horarios_empleados  ·  1:N con citas (como profesional)  ·  M:N con servicios via empleado_servicios",
        "atributos": [
            ("PK", "id_empleado",    "SERIAL",       "Clave primaria"),
            ("",   "nombre",         "VARCHAR(100)", "Nombre completo de la especialista"),
            ("",   "especialidad",   "VARCHAR(100)", "Área de especialización (ej: Nail Artist)"),
            ("",   "activo",         "BOOLEAN",      "Disponible para citas (TRUE) o desactivada"),
            ("",   "fecha_registro", "TIMESTAMP",    "Fecha de registro en el sistema"),
        ]
    },
    {
        "nombre": "SERVICIOS",
        "descripcion": "Catálogo completo de servicios ofrecidos por el salón con precios y duraciones.",
        "relaciones": "M:N con empleados via empleado_servicios  ·  1:N con citas",
        "atributos": [
            ("PK", "id_servicio",      "SERIAL",         "Clave primaria"),
            ("",   "nombre_servicio",  "VARCHAR(100)",    "Nombre del servicio (ej: Manicure Clásico)"),
            ("",   "descripcion",      "TEXT",            "Descripción del servicio"),
            ("",   "precio_total",     "NUMERIC(10,2)",   "Precio en pesos colombianos COP"),
            ("",   "duracion_minutos", "INTEGER",         "Duración estimada en minutos"),
            ("",   "activo",           "BOOLEAN",         "Publicado en el catálogo de agendamiento"),
        ]
    },
    {
        "nombre": "EMPLEADO_SERVICIOS",
        "descripcion": "Tabla pivote M:N que define qué servicios puede realizar cada especialista.",
        "relaciones": "N:1 con empleados  ·  N:1 con servicios",
        "atributos": [
            ("PK/FK", "id_empleado", "INTEGER → empleados", "Parte de la clave primaria compuesta"),
            ("PK/FK", "id_servicio", "INTEGER → servicios",  "Parte de la clave primaria compuesta"),
        ]
    },
    {
        "nombre": "HORARIOS_EMPLEADOS",
        "descripcion": "Configuración de horarios semanales por especialista. Define días y franjas de atención.",
        "relaciones": "N:1 con empleados — cascade delete si se elimina el empleado",
        "atributos": [
            ("PK", "id_horario",   "SERIAL",                 "Clave primaria"),
            ("FK", "id_empleado",  "INTEGER → empleados",    "Empleado al que pertenece este horario"),
            ("",   "dia_semana",   "INTEGER (0-6)",           "0=Dom, 1=Lun, 2=Mar, 3=Mié, 4=Jue, 5=Vie, 6=Sáb"),
            ("",   "hora_inicio",  "TIME",                   "Hora de inicio del turno laboral"),
            ("",   "hora_fin",     "TIME",                   "Hora de fin del turno laboral"),
        ]
    },
    {
        "nombre": "CITAS",
        "descripcion": "Reservas agendadas por los clientes. Tabla central del sistema con estados, abonos y tokens de gestión.",
        "relaciones": "N:1 con usuario (cliente)  ·  N:1 con empleados (profesional)  ·  N:1 con servicios  ·  1:1 con pagos",
        "atributos": [
            ("PK", "id_cita",           "SERIAL",                     "Clave primaria"),
            ("FK", "id_cliente",        "INTEGER → usuario",          "Cliente que agendó — CASCADE delete"),
            ("FK", "id_empleado",       "INTEGER → empleados",        "Especialista asignada — SET NULL si se elimina"),
            ("FK", "id_servicio",       "INTEGER → servicios",        "Servicio reservado — RESTRICT delete"),
            ("",   "fecha_hora_inicio", "TIMESTAMP",                  "Inicio de la cita"),
            ("",   "fecha_hora_fin",    "TIMESTAMP",                  "Fin calculado de la cita"),
            ("",   "monto_total",       "NUMERIC(10,2)",              "Precio total del servicio"),
            ("",   "monto_abono",       "NUMERIC(10,2)",              "Abono pagado (mínimo $5.000 COP)"),
            ("",   "saldo_pendiente",   "NUMERIC(10,2)",              "monto_total - monto_abono"),
            ("",   "estado",            "ENUM estado_cita_enum",      "pendiente_pago | confirmada | en_atencion | completada | cancelada | no_asistio"),
            ("",   "reembolsado",       "BOOLEAN",                    "TRUE si el abono fue devuelto"),
            ("UK", "codigo_reserva",    "VARCHAR(20)",                "Código único generado con secrets.token_urlsafe"),
            ("UK", "token_gestion",     "VARCHAR(32)",                "Token URL-safe para link de gestión por WhatsApp"),
            ("",   "notas",             "TEXT",                       "Observaciones y flags internos"),
            ("",   "fecha_creacion",    "TIMESTAMP",                  "Fecha de creación del registro"),
        ]
    },
    {
        "nombre": "PAGOS",
        "descripcion": "Transacciones de pago registradas para cada cita. Relación 1:1 con citas.",
        "relaciones": "1:1 con citas (UNIQUE en id_cita) — CASCADE delete si se elimina la cita",
        "atributos": [
            ("PK",    "id_pago",     "SERIAL",                   "Clave primaria"),
            ("FK/UK", "id_cita",     "INTEGER → citas",          "Cita asociada — una cita = un pago máximo"),
            ("",      "monto",       "NUMERIC(10,2)",            "Monto del pago registrado"),
            ("",      "metodo_pago", "ENUM metodo_pago_enum",    "efectivo | tarjeta | transferencia | nequi | daviplata"),
            ("",      "estado_pago", "VARCHAR(20)",              "pendiente | completado | reembolsado"),
            ("",      "referencia",  "VARCHAR(100)",             "Número de transacción o comprobante"),
            ("",      "fecha_pago",  "TIMESTAMP",                "Fecha y hora del pago"),
            ("",      "notas",       "TEXT",                     "Observaciones del pago"),
        ]
    },
    {
        "nombre": "NOTIFICACIONES",
        "descripcion": "Alertas internas del sistema para clientes y administradores. Incluye link de destino.",
        "relaciones": "N:1 con usuario — CASCADE delete si se elimina el usuario",
        "atributos": [
            ("PK", "id",         "SERIAL",              "Clave primaria"),
            ("FK", "id_usuario", "INTEGER → usuario",   "Destinatario de la notificación"),
            ("",   "titulo",     "VARCHAR(200)",        "Asunto o título de la notificación"),
            ("",   "mensaje",    "TEXT",                "Cuerpo del mensaje"),
            ("",   "target",     "VARCHAR(300)",        "URL Flask de redirección al hacer clic"),
            ("",   "leido",      "BOOLEAN",             "Estado de lectura (FALSE = no leído)"),
            ("",   "fecha",      "TIMESTAMP",           "Fecha y hora de creación"),
        ]
    },
    {
        "nombre": "AUDITORIA_USUARIOS",
        "descripcion": "Log inmutable de todas las acciones críticas sobre cuentas. Incluye snapshot de datos en el momento de la acción.",
        "relaciones": "N:1 con usuario (afectado) — SET NULL  ·  N:1 con usuario (actor/admin) — SET NULL",
        "atributos": [
            ("PK", "id",           "SERIAL",              "Clave primaria"),
            ("FK", "id_usuario",   "INTEGER → usuario",   "Usuario afectado — SET NULL al borrar"),
            ("FK", "id_actor",     "INTEGER → usuario",   "Admin que ejecutó la acción — SET NULL"),
            ("",   "nombre",       "VARCHAR(100)",        "Snapshot del nombre al momento"),
            ("",   "email",        "VARCHAR(120)",        "Snapshot del email al momento"),
            ("",   "telefono",     "VARCHAR(20)",         "Snapshot del teléfono al momento"),
            ("",   "tipo_usuario", "VARCHAR(20)",         "Snapshot del rol al momento"),
            ("",   "accion",       "VARCHAR(50)",         "login | logout | editar | eliminar | desactivar"),
            ("",   "detalle",      "TEXT",                "Descripción del cambio"),
            ("",   "ip_address",   "VARCHAR(45)",         "Dirección IP (IPv4 o IPv6)"),
            ("",   "fecha",        "TIMESTAMP",           "Fecha y hora del evento"),
        ]
    },
    {
        "nombre": "CONFIGURACIONES",
        "descripcion": "Parámetros operativos configurables del sistema (clave-valor). Permite ajustar comportamiento sin modificar código.",
        "relaciones": "N:1 con usuario (creador) — SET NULL  ·  N:1 con usuario (modificador) — SET NULL",
        "atributos": [
            ("PK", "id",                  "SERIAL",            "Clave primaria"),
            ("UK", "clave",               "VARCHAR(120)",      "Identificador del parámetro (ej: abono_minimo)"),
            ("",   "valor",               "TEXT",              "Valor en texto plano"),
            ("",   "descripcion",         "TEXT",              "Explicación del parámetro"),
            ("FK", "creado_por",          "INTEGER → usuario", "Admin que creó el parámetro — SET NULL"),
            ("FK", "modificado_por",      "INTEGER → usuario", "Admin que lo modificó por última vez — SET NULL"),
            ("",   "fecha_creacion",      "TIMESTAMP",         "Fecha de creación"),
            ("",   "fecha_actualizacion", "TIMESTAMP",         "Fecha de última modificación"),
        ]
    },
]

# ════════════════════════════════════════════════════════════════════════════
# CASOS DE PRUEBA
# ════════════════════════════════════════════════════════════════════════════
CASOS_PRUEBA = [
    {
        "modulo": "MÓDULO: AUTENTICACIÓN",
        "casos": [
            ("CP-001","Login exitoso como Administrador","Crítica","admin@rossmix.com / admin123","Redirige a /admin/ — Dashboard admin visible","PASA"),
            ("CP-002","Login exitoso como Cliente","Alta","andrea.vargas@email.com / cliente123","Redirige a /dashboard/cliente — Bienvenido/a visible","PASA"),
            ("CP-003","Login exitoso como Especialista","Alta","ana.rodriguez@rossmix.com / especialista123","Redirige a /especialista/dashboard","PASA"),
            ("CP-004","Login con contraseña incorrecta","Alta","admin@rossmix.com / wrongpass","Mensaje: 'Email o contraseña incorrectos. Intentos restantes: 9'","PASA"),
            ("CP-005","Bloqueo por fuerza bruta (10 intentos)","Crítica","Cualquier email, 10 intentos fallidos","Mensaje: 'Cuenta bloqueada 30 minutos'. No permite más intentos","PASA"),
            ("CP-006","Registro con email duplicado","Media","Email ya registrado","Mensaje: 'Este email ya está registrado'","PASA"),
            ("CP-007","Registro con teléfono inválido","Alta","Teléfono con menos de 10 dígitos","Mensaje de validación — no permite crear cuenta","PASA"),
            ("CP-008","Cierre de sesión exitoso","Media","Clic en 'Cerrar Sesión'","Redirige al inicio. Sesión destruida. Registro en auditoría","PASA"),
        ]
    },
    {
        "modulo": "MÓDULO: AGENDAMIENTO DE CITAS (4 PASOS)",
        "casos": [
            ("CP-009","Paso 1 — Seleccionar servicio activo","Alta","Clic en servicio del catálogo","Avanza al paso 2 con el servicio seleccionado","PASA"),
            ("CP-010","Paso 2 — Elegir especialista específica","Alta","Seleccionar Ana Rodríguez","Avanza al paso 3 con id_empleado=1","PASA"),
            ("CP-011","Paso 2 — Opción 'Cualquiera disponible'","Alta","Seleccionar opción aleatoria","Sistema elige empleado activo con disponibilidad","PASA"),
            ("CP-012","Paso 3 — Ver horarios disponibles","Crítica","Seleccionar fecha futura válida","API retorna slots de 30 min disponibles para esa fecha","PASA"),
            ("CP-013","Paso 3 — Fecha pasada bloqueada","Alta","Seleccionar fecha anterior a hoy","No se muestran slots. Validación impide avanzar","PASA"),
            ("CP-014","Paso 4 — Confirmar cita con abono","Crítica","Confirmar con datos válidos","Cita creada con estado pendiente_pago, codigo_reserva generado","PASA"),
            ("CP-015","Confirmar con empleado inactivo (manipulación)","Crítica","Enviar id_empleado de empleada inactiva","Error: 'La especialista seleccionada no está disponible'","PASA"),
            ("CP-016","Confirmar con duración manipulada","Crítica","Enviar fecha_hora_fin incorrecta","Sistema corrige fecha_hora_fin con duración real del servicio","PASA"),
            ("CP-017","Confirmar con menos de 30 min de anticipación","Alta","Cita para 15 minutos después","Error: 'La cita debe ser con al menos 30 minutos de anticipación'","PASA"),
        ]
    },
    {
        "modulo": "MÓDULO: GESTIÓN DE CITAS (CLIENTE)",
        "casos": [
            ("CP-018","Ver mis citas futuras","Media","Acceder a /citas/mis-citas","Lista citas pendientes/confirmadas con servicios y especialistas","PASA"),
            ("CP-019","Cancelar cita con ≥2h de anticipación","Crítica","Cancelar cita con 3h antes","Cita cancelada. Reembolso de $5.000 COP procesado. Notificación enviada","PASA"),
            ("CP-020","Cancelar cita con <2h de anticipación","Crítica","Cancelar cita con 1h antes","Cita cancelada SIN reembolso. Mensaje de política mostrado","PASA"),
            ("CP-021","Reprogramar cita con ≥2h anticipación","Crítica","Reprogramar a nueva fecha válida","Cita original cancelada. Nueva cita creada. Abono transferido","PASA"),
            ("CP-022","Reprogramar con <2h anticipación","Alta","Intentar reprogramar tarde","Error: 'Debes reprogramar con al menos 2 horas de anticipación'","PASA"),
            ("CP-023","Descargar comprobante PDF","Media","Clic en Descargar PDF","Archivo PDF con datos de la cita descargado","PASA"),
            ("CP-024","Reagendar cita No Asistió con crédito","Alta","Reagendar desde cita no_asistio","Crédito aplicado al nuevo agendamiento. Flag [CREDITO_CONSUMIDO] en BD","PASA"),
            ("CP-025","Intento de reutilizar crédito","Crítica","Intentar reagendar segunda vez con mismo crédito","Error: 'El crédito de esta cita ya fue utilizado'","PASA"),
        ]
    },
    {
        "modulo": "MÓDULO: PANEL ADMINISTRATIVO",
        "casos": [
            ("CP-026","Dashboard admin — estadísticas del día","Media","Acceder a /admin/","Muestra citas_hoy, total_clientes, ingresos_mes, pagos_pendientes","PASA"),
            ("CP-027","Agenda diaria — ver citas del día","Alta","Acceder a /admin/agenda-diaria","Cuadrícula de citas agrupadas por profesional para hoy","PASA"),
            ("CP-028","Cambiar estado cita a En Atención","Crítica","Marcar cliente como llegado","Estado cambia a en_atencion. Notificación al cliente","PASA"),
            ("CP-029","Registrar pago con monto excesivo","Crítica","Ingresar monto > saldo pendiente","Error: 'El monto no puede superar el saldo pendiente de $X'","PASA"),
            ("CP-030","Registrar pago correcto","Crítica","Monto = saldo pendiente","Pago registrado. Cita pasa a completada. Cliente notificado","PASA"),
            ("CP-031","Crear empleada sin servicios asignados","Alta","Guardar empleada sin checkboxes","Advertencia mostrada. Empleada creada pero no visible en agendamiento","PASA"),
            ("CP-032","Editar cliente con email inválido","Alta","Email sin @ en edición de cliente","Error de validación. No guarda. Mensaje específico del error","PASA"),
            ("CP-033","Eliminar cliente con citas futuras","Alta","Intentar eliminar cliente activo","Error: 'No se puede eliminar. El cliente tiene X cita(s) pendiente(s)'","PASA"),
            ("CP-034","Asignación batch de citas — validar disponibilidad","Crítica","Asignar dos citas al mismo empleado mismo horario","Segunda cita rechazada. Error de disponibilidad mostrado","PASA"),
            ("CP-035","Exportar pagos a Excel — período mensual","Media","Exportar pagos del mes","Archivo .xlsx descargado con todos los pagos del período","PASA"),
        ]
    },
    {
        "modulo": "MÓDULO: ESPECIALISTA",
        "casos": [
            ("CP-036","Ver citas disponibles para aceptar","Alta","Acceder a /especialista/citas-disponibles","Lista de citas sin asignar para sus servicios","PASA"),
            ("CP-037","Aceptar cita disponible","Crítica","Clic en aceptar cita","Cita asignada. Estado a confirmada. Cliente notificado con nombre de especialista","PASA"),
            ("CP-038","Aceptar cita ya tomada por otra","Crítica","Dos especialistas aceptan simultáneamente","Segunda recibe: 'Esta cita ya fue tomada por otra especialista'","PASA"),
            ("CP-039","Ver mis citas del día","Media","Acceder a mis citas","Lista de citas asignadas con horarios y datos del cliente","PASA"),
            ("CP-040","Especialista con empleado desactivado pierde acceso","Crítica","Admin desactiva el empleado vinculado","Próximo request: sesión cerrada. Mensaje de cuenta desactivada","PASA"),
        ]
    },
    {
        "modulo": "MÓDULO: SEGURIDAD",
        "casos": [
            ("CP-041","Acceso a /admin/ sin sesión","Crítica","GET /admin/ sin cookie de sesión","Redirige a /login. No muestra datos del panel","PASA"),
            ("CP-042","Acceso a /admin/ con sesión de cliente","Crítica","Cliente intenta /admin/","Redirige a dashboard cliente. No hay acceso","PASA"),
            ("CP-043","Token CSRF faltante en formulario","Crítica","Enviar POST sin csrf_token","HTTP 400: 'Falta el token CSRF'","PASA"),
            ("CP-044","XSS en nombre de cliente","Crítica","Nombre: <script>alert(1)</script>","Texto se muestra literal sin ejecutar. textContent usado","PASA"),
            ("CP-045","Manipulación de monto de pago","Crítica","Enviar monto > saldo vía POST directo","Error: monto limitado al saldo pendiente real","PASA"),
            ("CP-046","Solapamiento de citas simultáneas","Crítica","Dos requests simultáneos para mismo horario/empleado","Segunda request falla con UniqueViolation del índice parcial en BD","PASA"),
        ]
    },
    {
        "modulo": "MÓDULO: MÓVIL Y RESPONSIVE",
        "casos": [
            ("CP-047","Acceso desde celular en red WiFi","Alta","192.168.20.25:5000 desde Android","Página carga correctamente. Layout responsive activo","PASA"),
            ("CP-048","Menú hamburguesa se abre","Alta","Tocar ≡ en pantalla <768px","Menú desplegable aparece con opciones según rol","PASA"),
            ("CP-049","Menú hamburguesa se cierra al tocar fuera","Media","Tocar fuera del menú abierto","Menú se oculta. Ícono vuelve a ≡","PASA"),
            ("CP-050","Formularios no hacen zoom en iOS","Media","Enfocar input en iPhone","No hay zoom automático (font-size ≥ 16px en todos los campos)","PASA"),
        ]
    },
]

# ════════════════════════════════════════════════════════════════════════════
# ERRORES DEL PROCESO (resumen compacto)
# ════════════════════════════════════════════════════════════════════════════
ERRORES_PROCESO = [
    # (id, fase, severidad, categoria, descripcion_corta, archivo, correccion_corta, commit)
    ("E-001","Revisión Inicial","CRÍTICO","Seguridad","credenciales_login.md con contraseñas en texto plano en GitHub","credenciales_login.md","Eliminado del repo. Agregado al .gitignore","49e116c"),
    ("E-002","Revisión Inicial","CRÍTICO","Bug","token_gestion nunca generado — funcionalidad de gestión rota","citas_service.py","secrets.token_urlsafe(24) en crear_cita()","49e116c"),
    ("E-003","Revisión Inicial","CRÍTICO","Lógica","reagendar_no_asistio prometía crédito sin implementarlo","views/citas.py","Crédito en BD con flag [CREDITO_CONSUMIDO]. GET→POST","49e116c"),
    ("E-004","Revisión Inicial","ALTO","Funcionalidad","reprogramar_cita_form sin ruta POST — reprogramación incompleta","views/citas.py","Creada ruta POST completa con transferencia de abono","49e116c"),
    ("E-005","Revisión Inicial","ALTO","Bug","reportes_service usa Usuario en lugar de Empleado para horarios","services/reportes_service.py","Reescrito con JOINs. Bug de modelo corregido","49e116c"),
    ("E-006","Revisión Inicial","ALTO","Inconsistencia","Lógica de abono diferente entre admin y cliente","admin/pagos.py","Unificado: ambos acumulan el abono","49e116c"),
    ("E-007","Revisión Inicial","ALTO","Bug","Lógica de disponibilidad duplicada en vista y servicio","views/citas.py","Vista delega completamente al servicio","49e116c"),
    ("E-008","Revisión Inicial","ALTO","Bug","ENUMs con create_constraint=False rompían en BD nueva","models/cita.py","Cambiado a create_type=False (correcto para psycopg3)","49e116c"),
    ("E-009","Revisión Inicial","MEDIO","Seguridad","admin_required no verificaba usuario.activo en BD","utils/decorators.py","Verifica activo en BD en cada request","49e116c"),
    ("E-010","Revisión Inicial","MEDIO","Sesión","Sesión sin timeout configurado","app/config.py","PERMANENT_SESSION_LIFETIME = 8 horas","49e116c"),
    ("E-011","Seguridad","CRÍTICO","Seguridad","Campos ocultos de cita sin validación — empleado manipulable","views/citas.py:176","Validación completa: activo, servicio, duración, disponibilidad","32cbe74"),
    ("E-012","Seguridad","CRÍTICO","Seguridad","Pagos con importe arbitrario — sobrepagos posibles","views/citas.py:383","Monto limitado al saldo_pendiente real","32cbe74"),
    ("E-013","Seguridad","CRÍTICO","Seguridad","Crédito de inasistencia reutilizable múltiples veces","views/citas.py:545","Flag en BD. GET→POST. Verificación antes de aplicar","32cbe74"),
    ("E-014","Seguridad","CRÍTICO","Seguridad","Reasignación de citas sin validar empleados activos","admin/citas.py:108","Valida activo + servicio + disponibilidad","32cbe74"),
    ("E-015","Seguridad","CRÍTICO","Seguridad","Condición de carrera en reservas simultáneas","citas_service.py:15","Índice único parcial en PostgreSQL","32cbe74"),
    ("E-016","Seguridad","CRÍTICO","Seguridad","XSS almacenado en admin via innerHTML con datos de usuario","admin/empleados.html:454","innerHTML→textContent/createElement","32cbe74"),
    ("E-017","Seguridad","MEDIO","Seguridad","Especialistas con empleados desactivados conservaban acceso","decorators.py:68","Verifica empleado.activo en cada request","32cbe74"),
    ("E-018","Seguridad","MEDIO","Seguridad","Sin protección contra fuerza bruta en login","views/auth.py:23","Rate limiting: 10 intentos/15min, bloqueo 30min","32cbe74"),
    ("E-019","Seguridad","MEDIO","Config","debug=True hardcodeado en app.py y run.py","app.py, run.py","debug dinámico según FLASK_ENV","32cbe74"),
    ("E-020","Seguridad","MEDIO","Config","SECRET_KEY cambiaba sin advertencia — sesiones inválidas","config.py:45","sys.exit(1) en producción si falta SECRET_KEY","32cbe74"),
    ("E-021","Seguridad","MEDIO","Bug","reprogramar.html llamaba endpoint /TOKEN/confirmar inexistente","templates/reprogramar.html:280","Corregido a POST /citas/reprogramar/<id_cita>","32cbe74"),
    ("E-022","Seguridad","MEDIO","Consistencia","Mezcla datetime.now() y utcnow() en todos los modelos","Múltiples modelos","Unificado a datetime.now() (28 archivos)","32cbe74"),
    ("E-023","Seguridad","MEDIO","Datos","Email duplicado en crear_usuarios.py","scripts/crear_usuarios.py","Email corregido","32cbe74"),
    ("E-024","Seguridad","BAJO","Seguridad","POSTGRES_PASSWORD=1234 hardcodeado en docker-compose","docker-compose.yml","Cambiado a variable requerida ${DB_PASSWORD}","32cbe74"),
    ("E-025","Calidad","CRÍTICO","Bug","session['email'] nunca guardado — dropdown siempre vacío","views/auth.py","session['email'] = usuario.email en login","84920d8"),
    ("E-026","Calidad","CRÍTICO","Bug","Race condition en asignación aleatoria de empleado","views/citas.py","Un empleado elegido antes de consultar slots","84920d8"),
    ("E-027","Calidad","ALTO","Performance","N+1 queries en listado de empleados","admin/empleados.py","GROUP BY en una sola query","84920d8"),
    ("E-028","Calidad","ALTO","Performance","N+1 queries en listado de clientes","admin/clientes.py","GROUP BY en una sola query","84920d8"),
    ("E-029","Calidad","ALTO","Seguridad","Edición cliente sin validación email/teléfono","admin/clientes.py","Validación regex antes de guardar","84920d8"),
    ("E-030","Calidad","ALTO","Performance","citas_completadas cargaba TODAS las citas en memoria","models/usuario.py","COUNT en BD","84920d8"),
    ("E-031","Calidad","ALTO","Seguridad","flash(str(e)) exponía internos de SQLAlchemy","views/citas.py","Mensajes genéricos seguros","84920d8"),
    ("E-032","Calidad","ALTO","Bug","reprogramar_cita no filtraba empleados activos","views/citas.py:701","JOIN con Empleado.activo == True","84920d8"),
    ("E-033","Calidad","ALTO","Bug","30+ ocurrencias query.get_or_404() deprecado SQLAlchemy 2.x","10 archivos","31 reemplazos por db.get_or_404()","84920d8"),
    ("E-034","Calidad","ALTO","Seguridad","Sin rate limiting en /registro","views/auth.py","Mismo mecanismo que login","84920d8"),
    ("E-035","Calidad","ALTO","Seguridad","Cookies sin SECURE/HTTPONLY/SAMESITE","config.py","Agregadas las 3 flags","84920d8"),
    ("E-036","Calidad","ALTO","Seguridad","Código reserva con random no criptográfico","citas_service.py","secrets.token_urlsafe(6)[:8]","84920d8"),
    ("E-037","Calidad","MEDIO","Bug","add_notificacion hacía commit propio — rompía atomicidad","utils/helpers.py","Eliminado commit. Llamador es responsable","84920d8"),
    ("E-038","Calidad","MEDIO","Seguridad","citas_asignar_batch sin validar disponibilidad","admin/citas.py","Valida activo + CitaService.validar_disponibilidad","84920d8"),
    ("E-039","Calidad","BAJO","Performance","Índices faltantes en columnas frecuentes","models/usuario.py, cita.py","index=True en tipo_usuario, estado, fecha_hora_inicio","84920d8"),
    ("E-040","UI/UX","ALTO","UI","Modales con headers morado/verde fuera de paleta","admin/empleados.html, servicios.html","Cambiado a gradiente rosa Rossmix","913c2b3"),
    ("E-041","UI/UX","ALTO","UI","Filas horarios con fondo rosa — texto ilegible","admin/horarios.html","Restauradas filas blancas, badge rosa en esquina","5757b09"),
    ("E-042","UI/UX","ALTO","UI","Botón Crear Cuenta con gradiente dorado/marrón","registro.html","Gradiente rosa Rossmix","f20f46d"),
    ("E-043","UI/UX","MEDIO","UI","Secciones index.html con 8 colores distintos","templates/index.html","Todos a variaciones de rosa Rossmix","f20f46d"),
    ("E-044","UI/UX","MEDIO","UI","dashboard_admin con 37 colores fuera de paleta","dashboard_admin.html","37 reemplazos a paleta Rossmix","56c171a"),
    ("E-045","UI/UX","MEDIO","UI","526 colores fuera de paleta en 29 archivos","29 templates","Script automatizado: 526 reemplazos","4523a4f"),
    ("E-046","UI/UX","BAJO","UI","Textos blancos invisibles sobre fondos claros","dashboard_*.html","Cambiados a #444/#555 para legibilidad","4523a4f"),
    ("E-047","Infra","ALTO","Infra","Flask solo en 127.0.0.1 — inaccesible desde celular","app.py","host='0.0.0.0'. Firewall abierto","03f35da"),
    ("E-048","Infra","ALTO","Infra",".env con BOM UTF-8 — python-dotenv no leía variables",".env","Reescrito sin BOM","03f35da"),
    ("E-049","Infra","ALTO","Infra","docker-compose con host.docker.internal no resolvía","docker-compose.yml","Cambiado a localhost / db (servicio Docker)","b769896"),
    ("E-050","Infra","MEDIO","Infra","Dockerfile con Python 3.11 — proyecto requiere 3.13","Dockerfile","Multi-stage build con python:3.13-slim","b769896"),
    ("E-051","Infra","MEDIO","Infra","YAML inválido — :? rompe parser de Docker Compose","docker-compose.yml","Cambiado a ${DB_PASSWORD}","b769896"),
    ("E-052","Infra","MEDIO","UX","Menú hamburguesa no funcionaba en móvil","templates/base.html","IIFE independiente. position:fixed. stopPropagation()","6796dc2"),
    ("E-053","Infra","BAJO","Git","cloudflared.exe (73MB) sin .gitignore","cloudflared.exe","Agregado al .gitignore","84920d8"),
    ("E-054","Infra","BAJO","Git","Dos entornos virtuales con rutas rotas",".venv/, .venv313/","Eliminados y recreado .venv limpio","manual"),
    ("E-055","Formularios","CRÍTICO","Seguridad","CSRF faltante en todos los formularios admin","admin/*.html x5","Token CSRF + interceptor fetch global","2307d3c"),
    ("E-056","Formularios","ALTO","Bug","Campo monto con step=1000 rechazaba valores válidos","admin/pagos_form.html","step='1' min='5000'. Spinner oculto","2307d3c"),
    ("E-057","Formularios","ALTO","Bug","CSS sobreescribía ocultamiento de navbar en móvil","app/static/style.css","display:none !important garantizado","f811e75"),
    ("E-058","Formularios","MEDIO","UX","Login y Registro mostraban todos los links juntos","login.html, registro.html","auth_base.html con navbar mínima por rol","2307d3c"),
    ("E-059","Formularios","MEDIO","UX","'Pagos por Confirmar' siempre visible aunque fuera 0","base.html, helpers.py","Calculado en context processor. Oculto si es 0","2307d3c"),
    ("E-060","Formularios","MEDIO","UX","Botones acción con colores inconsistentes","app/static/style.css","Todos unificados a paleta #fff0f6 / #c41e3a","56c171a"),
]

# ════════════════════════════════════════════════════════════════════════════
# SOFTWARE USADO
# ════════════════════════════════════════════════════════════════════════════
SOFTWARE = [
    ("Backend / Framework", "Flask", "2.3.3", "Microframework web Python para API REST y rutas"),
    ("Backend / Framework", "Flask-SQLAlchemy", "3.1.1", "ORM para PostgreSQL — modelos y queries"),
    ("Backend / Framework", "SQLAlchemy", "2.0.36", "Core ORM anclado para evitar breaking changes"),
    ("Backend / Framework", "Werkzeug", "2.3.7", "Seguridad de contraseñas y middleware WSGI"),
    ("Base de Datos", "PostgreSQL", "18.x", "RDBMS principal. ENUMs, índices, vistas SQL"),
    ("Base de Datos", "psycopg[binary]", "3.3.4", "Driver PostgreSQL para Python 3 (psycopg3)"),
    ("Seguridad", "Flask-WTF", "1.1.1", "Protección CSRF en formularios"),
    ("Seguridad", "email-validator", "2.3.0", "Validación de formato de correo electrónico"),
    ("Autenticación", "Flask-Caching", "2.4.1", "Rate limiting de login vía SimpleCache"),
    ("Correo", "Flask-Mail", "0.10.0", "Envío de correos de confirmación y notificación"),
    ("Archivos", "openpyxl", "3.1.2", "Generación de reportes Excel (.xlsx)"),
    ("Archivos", "reportlab", "4.2.5", "Generación de comprobantes PDF de citas"),
    ("Archivos", "python-docx", "1.1.2", "Generación de documentación Word (.docx)"),
    ("Configuración", "python-dotenv", "1.0.1", "Lectura de variables de entorno desde .env"),
    ("Utilidades", "python-dateutil", "2.9.0", "Cálculo de fechas relativas (relativedelta)"),
    ("Servidor", "Gunicorn", "21.2.0", "Servidor WSGI de producción con workers"),
    ("Contenedores", "Docker", "29.7.2", "Contenedorización de la app y BD"),
    ("Contenedores", "Docker Compose", "v5.4.0", "Orquestación de servicios (web + db + seed)"),
    ("Base de Datos Docker", "PostgreSQL Alpine", "16-alpine", "Imagen ligera de PostgreSQL en contenedor"),
    ("Control de Versiones", "Git", "2.x", "Control de versiones del proyecto"),
    ("Repositorio", "GitHub", "—", "Repositorio: SaraKerrigan2001/Rossmix_flask"),
    ("Despliegue", "Render.com", "—", "Plataforma de hosting. render.yaml configurado"),
    ("Túnel Local", "Cloudflare Tunnel", "2026.8.2", "Exposición del servidor local a internet"),
    ("Lenguaje", "Python", "3.14.6", "Lenguaje principal del backend"),
    ("Frontend", "Jinja2", "3.1.6", "Motor de plantillas HTML para Flask"),
    ("Frontend", "Bootstrap Icons", "1.11.3", "Librería de iconos SVG"),
    ("Frontend", "Bootstrap CSS", "5.3.2", "Framework CSS responsivo"),
    ("Fuentes", "Google Fonts", "—", "Cormorant Garamond, Playfair Display, Inter"),
    ("Editor", "Kiro IDE", "—", "IDE con asistente IA para desarrollo"),
    ("BD Admin", "pgAdmin", "—", "Administración visual de PostgreSQL"),
]


# ════════════════════════════════════════════════════════════════════════════
# GENERAR EXCEL
# ════════════════════════════════════════════════════════════════════════════
def generar_excel():
    wb = openpyxl.Workbook()

    # ── HOJA 1: PORTADA ───────────────────────────────────────────────────────
    ws0 = wb.active
    ws0.title = 'Portada'
    ws0.sheet_view.showGridLines = False
    ws0.column_dimensions['A'].width = 5
    ws0.column_dimensions['B'].width = 55

    ws0.row_dimensions[3].height = 50
    t = ws0.cell(row=3, column=2, value='ROSSMIX — SALÓN DE BELLEZA')
    t.font = Font(bold=True, size=28, color=R, name='Calibri')
    t.alignment = align('left', 'center')

    ws0.row_dimensions[4].height = 22
    s = ws0.cell(row=4, column=2, value='Sistema de Agendamiento de Citas — Documentación Técnica')
    s.font = Font(size=13, color=GRIS, name='Calibri')
    s.alignment = align('left', 'center')

    ws0.row_dimensions[5].height = 18
    f = ws0.cell(row=5, column=2, value=f'Versión 2.0  ·  {date.today().strftime("%d/%m/%Y")}')
    f.font = Font(size=10, color='9CA3AF', italic=True, name='Calibri')
    f.alignment = align('left', 'center')

    hojas_info = [
        ('Diseño BD',    'Modelo relacional con 10 tablas: atributos, tipos, PK/FK y relaciones'),
        ('Casos Prueba', '50 casos de prueba organizados por módulo con resultado y estado'),
        ('Informe Errores', '60 errores corregidos durante el proceso de desarrollo'),
        ('Software Usado', '30 herramientas y tecnologías del proyecto'),
    ]
    r = 8
    merge_title(ws0, r, 2, 3, 'CONTENIDO DEL DOCUMENTO', R, W, 11)
    r += 1
    for hoja, desc in hojas_info:
        ws0.row_dimensions[r].height = 20
        h = ws0.cell(row=r, column=2, value=hoja)
        h.font = font(bold=True, size=10, color=R)
        h.fill = fill(S)
        h.alignment = align('left', 'center')
        h.border = border()
        d = ws0.cell(row=r, column=3, value=desc)
        d.font = font(size=10)
        d.fill = fill(W)
        d.alignment = align('left', 'center', wrap=True)
        d.border = border()
        r += 1
    ws0.column_dimensions['C'].width = 60

    # ── HOJA 2: DISEÑO BD ─────────────────────────────────────────────────────
    ws_bd = wb.create_sheet('Diseño BD')
    ws_bd.sheet_view.showGridLines = False
    ws_bd.column_dimensions['A'].width = 20
    ws_bd.column_dimensions['B'].width = 25
    ws_bd.column_dimensions['C'].width = 22
    ws_bd.column_dimensions['D'].width = 50

    # Título grande
    ws_bd.row_dimensions[1].height = 40
    ws_bd.merge_cells('A1:D1')
    tt = ws_bd.cell(row=1, column=1, value='DISEÑO DE LA BASE DE DATOS  —  TABLA')
    tt.font = Font(bold=True, size=20, color=D, name='Calibri')
    tt.alignment = align('center', 'center')

    ws_bd.row_dimensions[2].height = 15
    ws_bd.row_dimensions[3].height = 18
    sub = ws_bd.cell(row=3, column=1, value=f'Rossmix — Sistema de Agendamiento de Citas  ·  {date.today().strftime("%d/%m/%Y")}')
    sub.font = Font(italic=True, size=10, color=GRIS, name='Calibri')
    sub.alignment = align('center', 'center')
    ws_bd.merge_cells('A3:D3')

    row = 5
    for tabla in TABLAS:
        # NOMBRE
        ws_bd.row_dimensions[row].height = 20
        ws_bd.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        nc = ws_bd.cell(row=row, column=1, value='NOMBRE')
        nc.fill = fill(R); nc.font = font(bold=True, size=10, color=W)
        nc.alignment = align('left', 'center'); nc.border = border()
        nv = ws_bd.cell(row=row, column=2, value=tabla['nombre'])
        nv.font = font(bold=True, size=11, color=D)
        nv.fill = fill(G); nv.alignment = align('left', 'center'); nv.border = border()
        row += 1

        # DESCRIPCIÓN
        ws_bd.row_dimensions[row].height = 35
        ws_bd.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        dc = ws_bd.cell(row=row, column=1, value='DESCRIPCIÓN')
        dc.fill = fill(R); dc.font = font(bold=True, size=9, color=W)
        dc.alignment = align('left', 'center'); dc.border = border()
        dv = ws_bd.cell(row=row, column=2, value=tabla['descripcion'])
        dv.font = font(size=9, color=D)
        dv.fill = fill(W); dv.alignment = align('left', 'top', wrap=True); dv.border = border()
        row += 1

        # RELACIONES
        ws_bd.row_dimensions[row].height = 28
        ws_bd.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        rc = ws_bd.cell(row=row, column=1, value='RELACIONES')
        rc.fill = fill(R); rc.font = font(bold=True, size=9, color=W)
        rc.alignment = align('left', 'center'); rc.border = border()
        rv = ws_bd.cell(row=row, column=2, value=tabla['relaciones'])
        rv.font = font(size=9, color='1D4ED8', italic=True)
        rv.fill = fill(W); rv.alignment = align('left', 'top', wrap=True); rv.border = border()
        row += 1

        # ATRIBUTOS — header
        ws_bd.row_dimensions[row].height = 18
        for ci, hdr_txt in enumerate(['ATRIBUTOS','NOMBRE DEL ATRIBUTO','TIPO','DESCRIPCIÓN'],1):
            hc = ws_bd.cell(row=row, column=ci, value=hdr_txt)
            hc.fill = fill('374151')
            hc.font = font(bold=True, size=9, color=W)
            hc.alignment = align('center','center')
            hc.border = border('D1D5DB')
        row += 1

        for attr in tabla['atributos']:
            ws_bd.row_dimensions[row].height = 18
            clave, nombre, tipo, desc = attr
            # Clave
            kc = ws_bd.cell(row=row, column=1, value=clave)
            kc.font = font(bold=bool(clave), size=9, color='C41E3A' if 'PK' in clave else ('1D4ED8' if 'FK' in clave else D))
            kc.fill = fill(S if clave else W)
            kc.alignment = align('center','center')
            kc.border = border()
            # Nombre
            nm = ws_bd.cell(row=row, column=2, value=nombre)
            nm.font = font(bold='PK' in clave, size=9,
                           color='C41E3A' if 'PK' in clave else ('1D4ED8' if 'FK' in clave else D),
                           italic=('FK' in clave))
            nm.fill = fill(W); nm.alignment = align('left','center'); nm.border = border()
            # Tipo
            tp = ws_bd.cell(row=row, column=3, value=tipo)
            tp.font = font(size=9, color='374151')
            tp.fill = fill(W); tp.alignment = align('left','center'); tp.border = border()
            # Descripción
            dc2 = ws_bd.cell(row=row, column=4, value=f'• {desc}')
            dc2.font = font(size=9); dc2.fill = fill(W)
            dc2.alignment = align('left','center',wrap=True); dc2.border = border()
            row += 1

        row += 2  # espacio entre tablas

    # ── HOJA 3: CASOS DE PRUEBA ───────────────────────────────────────────────
    ws_cp = wb.create_sheet('Casos Prueba')
    ws_cp.sheet_view.showGridLines = False

    # Título
    ws_cp.merge_cells('A1:H1')
    ws_cp.row_dimensions[1].height = 30
    t2 = ws_cp.cell(row=1, column=1, value=f'CASOS DE PRUEBA  —  Rossmix  |  Versión Final  |  {date.today().strftime("%d/%m/%Y")}')
    t2.fill = fill(R); t2.font = font(bold=True, size=13, color=W)
    t2.alignment = align('center','center')

    # Subtítulo
    ws_cp.merge_cells('A2:H2')
    ws_cp.row_dimensions[2].height = 16
    ley = ws_cp.cell(row=2, column=1, value='Leyenda: Prioridad — Crítica (rojo)  |  Alta (naranja)  |  Media (azul)  |  Baja (verde)')
    ley.font = Font(italic=True, size=9, color=GRIS, name='Calibri')
    ley.alignment = align('center','center')

    cols_cp = ['ID','Descripción','Prioridad','Datos de Entrada','Resultado Esperado','Estado']
    widths_cp = [9, 45, 12, 40, 40, 10]
    for i, (col, w) in enumerate(zip(cols_cp, widths_cp), 1):
        ws_cp.column_dimensions[get_column_letter(i)].width = w

    PRIO_FILLS = {'Crítica':'FEE2E2','Alta':'FEF3C7','Media':'DBEAFE','Baja':'DCFCE7'}
    PRIO_FONTS = {'Crítica':'DC2626','Alta':'D97706','Media':'2563EB','Baja':'16A34A'}
    STATE_FILLS = {'PASA':'DCFCE7','PARCIAL':'FEF3C7','FALLA':'FEE2E2'}
    STATE_FONTS = {'PASA':'166534','PARCIAL':'92400E','FALLA':'991B1B'}
    STATE_EMO   = {'PASA':'✅ PASA','PARCIAL':'⚠️ PARCIAL','FALLA':'❌ FALLA'}

    row = 4
    for modulo in CASOS_PRUEBA:
        # Módulo header
        ws_cp.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws_cp.row_dimensions[row].height = 22
        mc = ws_cp.cell(row=row, column=1, value=modulo['modulo'])
        mc.fill = fill('1F2937'); mc.font = font(bold=True, size=10, color=W)
        mc.alignment = align('left','center')
        row += 1

        # Columna headers
        ws_cp.row_dimensions[row].height = 20
        for ci, h in enumerate(cols_cp, 1):
            hc = ws_cp.cell(row=row, column=ci, value=h)
            hc.fill = fill(R); hc.font = font(bold=True, size=9, color=W)
            hc.alignment = align('center','center')
            hc.border = border('D1D5DB')
        row += 1

        for caso in modulo['casos']:
            ws_cp.row_dimensions[row].height = 45
            cp_id, desc, prio, entrada, resultado, estado = caso
            bg = PRIO_FILLS.get(prio, W)

            data_cell(ws_cp, row, 1, cp_id, bg, R, bold=True, h='center', size=9)
            data_cell(ws_cp, row, 2, desc, W, D, size=9)
            # Prioridad con color
            pc = ws_cp.cell(row=row, column=3, value=prio)
            pc.fill = fill(bg); pc.font = font(bold=True, size=9, color=PRIO_FONTS.get(prio,D))
            pc.alignment = align('center','center'); pc.border = border()
            data_cell(ws_cp, row, 4, entrada, W, GRIS, size=9)
            data_cell(ws_cp, row, 5, resultado, W, D, size=9)
            # Estado
            sc = ws_cp.cell(row=row, column=6, value=STATE_EMO.get(estado, estado))
            sc.fill = fill(STATE_FILLS.get(estado, W))
            sc.font = font(bold=True, size=9, color=STATE_FONTS.get(estado, D))
            sc.alignment = align('center','center'); sc.border = border()
            row += 1

        # Resumen del módulo
        pasa = sum(1 for c in modulo['casos'] if c[5]=='PASA')
        total = len(modulo['casos'])
        ws_cp.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws_cp.row_dimensions[row].height = 16
        sm = ws_cp.cell(row=row, column=1,
                        value=f'  Subtotal: {pasa} PASAN ✅  |  {total-pasa} PARCIAL/FALLA  |  TOTAL: {total}  |  Cobertura: {int(pasa/total*100)}%')
        sm.fill = fill('F9FAFB'); sm.font = Font(italic=True, size=9, color=GRIS, name='Calibri')
        sm.alignment = align('right','center')
        row += 2

    # Resumen final
    total_cp = sum(len(m['casos']) for m in CASOS_PRUEBA)
    total_pasa = sum(1 for m in CASOS_PRUEBA for c in m['casos'] if c[5]=='PASA')
    ws_cp.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
    ws_cp.row_dimensions[row].height = 22
    rf = ws_cp.cell(row=row, column=1,
                    value=f'RESUMEN FINAL: {total_pasa} PASAN ✅  |  {total_cp-total_pasa} PARCIAL  |  TOTAL: {total_cp} CASOS  |  COBERTURA: {int(total_pasa/total_cp*100)}%')
    rf.fill = fill(R); rf.font = font(bold=True, size=11, color=W)
    rf.alignment = align('center','center')
    ws_cp.freeze_panes = 'A4'

    # ── HOJA 4: INFORME ERRORES ───────────────────────────────────────────────
    ws_err = wb.create_sheet('Informe Errores')
    ws_err.sheet_view.showGridLines = False

    ws_err.merge_cells('A1:H1')
    ws_err.row_dimensions[1].height = 30
    te = ws_err.cell(row=1, column=1, value='INFORME DE ERRORES Y CORRECCIONES  —  Rossmix')
    te.fill = fill(R); te.font = font(bold=True, size=13, color=W); te.alignment = align('center','center')

    ws_err.merge_cells('A2:H2')
    ws_err.row_dimensions[2].height = 16
    se = ws_err.cell(row=2, column=1, value=f'Total: 60 errores corregidos  ·  {date.today().strftime("%d/%m/%Y")}')
    se.font = Font(italic=True, size=9, color=GRIS, name='Calibri'); se.alignment = align('center','center')

    cols_err = ['ID','Fase','Severidad','Categoría','Descripción','Archivo','Corrección','Commit']
    widths_err = [8, 15, 10, 12, 48, 30, 48, 10]
    for i, w in enumerate(widths_err, 1):
        ws_err.column_dimensions[get_column_letter(i)].width = w

    ws_err.row_dimensions[3].height = 20
    for ci, h in enumerate(cols_err, 1):
        hc = ws_err.cell(row=3, column=ci, value=h)
        hc.fill = fill(R); hc.font = font(bold=True, size=9, color=W)
        hc.alignment = align('center','center'); hc.border = border()

    SEV_F = {'CRÍTICO':'FEE2E2','ALTO':'FEF3C7','MEDIO':'FFF0F6','BAJO':'F5F5F5'}
    SEV_C = {'CRÍTICO':'DC2626','ALTO':'D97706','MEDIO':'C41E3A','BAJO':'6B7280'}

    for i, err in enumerate(ERRORES_PROCESO, 4):
        ws_err.row_dimensions[i].height = 35
        eid, fase, sev, cat, desc, arch, corr, commit = err
        bg = SEV_F.get(sev, W)
        fill_alt = G if i % 2 == 0 else W
        data_cell(ws_err, i, 1, eid, bg, SEV_C.get(sev,D), bold=True, h='center')
        data_cell(ws_err, i, 2, fase, fill_alt, GRIS, size=9)
        sc2 = ws_err.cell(row=i, column=3, value=sev)
        sc2.fill = fill(bg); sc2.font = font(bold=True, size=9, color=SEV_C.get(sev,D))
        sc2.alignment = align('center','center'); sc2.border = border()
        data_cell(ws_err, i, 4, cat, fill_alt, GRIS, size=9)
        data_cell(ws_err, i, 5, desc, W, D, size=9)
        data_cell(ws_err, i, 6, arch, fill_alt, GRIS, size=8)
        data_cell(ws_err, i, 7, corr, W, D, size=9)
        data_cell(ws_err, i, 8, commit, fill_alt, GRIS, h='center', size=8)

    ws_err.freeze_panes = 'A4'

    # ── HOJA 5: SOFTWARE USADO ────────────────────────────────────────────────
    ws_sw = wb.create_sheet('Software Usado')
    ws_sw.sheet_view.showGridLines = False

    ws_sw.merge_cells('A1:E1')
    ws_sw.row_dimensions[1].height = 50
    ts = ws_sw.cell(row=1, column=1, value='SOFTWARE USADO')
    ts.font = Font(bold=True, size=32, color=D, name='Calibri')
    ts.alignment = align('center','center')

    ws_sw.merge_cells('A2:E2')
    ws_sw.row_dimensions[2].height = 20
    ss = ws_sw.cell(row=2, column=1, value=f'Tecnologías y herramientas del proyecto Rossmix  ·  {date.today().strftime("%d/%m/%Y")}')
    ss.font = Font(italic=True, size=11, color=GRIS, name='Calibri')
    ss.alignment = align('center','center')

    cols_sw = ['Categoría','Herramienta / Librería','Versión','Propósito / Uso']
    widths_sw = [22, 25, 12, 55]
    for i, w in enumerate(widths_sw, 1):
        ws_sw.column_dimensions[get_column_letter(i)].width = w

    ws_sw.row_dimensions[4].height = 20
    for ci, h in enumerate(cols_sw, 1):
        hc = ws_sw.cell(row=4, column=ci, value=h)
        hc.fill = fill(R); hc.font = font(bold=True, size=10, color=W)
        hc.alignment = align('center','center'); hc.border = border()

    cat_actual = ''
    row = 5
    for cat, tool, ver, uso in SOFTWARE:
        ws_sw.row_dimensions[row].height = 22
        bg = S if cat != cat_actual else W
        cat_actual = cat
        data_cell(ws_sw, row, 1, cat, S, R, bold=True, h='left', wrap=False)
        tc = ws_sw.cell(row=row, column=2, value=tool)
        tc.fill = fill(W); tc.font = font(bold=True, size=10, color=D)
        tc.alignment = align('left','center'); tc.border = border()
        data_cell(ws_sw, row, 3, ver, W, GRIS, h='center', size=9)
        data_cell(ws_sw, row, 4, uso, W, D, size=9, wrap=True)
        row += 1

    ws_sw.freeze_panes = 'A5'

    # ── GUARDAR ───────────────────────────────────────────────────────────────
    os.makedirs(os.path.join(BASE, 'docs'), exist_ok=True)
    out = os.path.join(BASE, 'docs', 'Rossmix_Documentacion_Completa.xlsx')
    wb.save(out)
    print(f'Excel guardado: {out}')
    print(f'Tamaño: {os.path.getsize(out)/1024:.1f} KB')
    print(f'Hojas: {len(wb.sheetnames)} — {wb.sheetnames}')

if __name__ == '__main__':
    generar_excel()
