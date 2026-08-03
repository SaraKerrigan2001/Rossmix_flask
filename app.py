from functools import wraps
from flask import Flask, render_template, request, redirect, url_for, flash, session, jsonify, send_file
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, timedelta, date, time
from decimal import Decimal
import random
import string
import io
import openpyxl
import logging
import uuid
import secrets
import os
from dataclasses import dataclass, field
from enum import Enum
from typing import Optional, Dict, Any
from dotenv import load_dotenv

# Cargar variables de entorno desde .env (desarrollo local)
load_dotenv()

# Configuración de logs profesional
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] (%(filename)s): %(message)s"
)

# PDF generation
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT

import os
app = Flask(
    __name__,
    template_folder=os.path.join(os.path.dirname(__file__), 'app', 'templates'),
    static_folder=os.path.join(os.path.dirname(__file__), 'app', 'static')
)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY') or os.urandom(32).hex()

# Configuración de PostgreSQL — leída desde variables de entorno / .env
_db_url = os.environ.get('DATABASE_URL') or (
    f"postgresql+psycopg://{os.environ.get('DB_USER','postgres')}:"
    f"{os.environ.get('DB_PASSWORD','1234')}@"
    f"{os.environ.get('DB_HOST','localhost')}:"
    f"{os.environ.get('DB_PORT','5432')}/"
    f"{os.environ.get('DB_NAME','Rossmix')}"
)
app.config['SQLALCHEMY_DATABASE_URI'] = _db_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

# ============================================================================
# SERVICIO DE PASARELA DE PAGOS — Reembolsos automáticos
# ============================================================================

class PasarelaPagoService:
    """
    Servicio encargado de gestionar reembolsos.
    En producción conecta con MercadoPago, Wompi, Stripe, etc.
    """
    REEMBOLSO_MONTO = 5000.0   # COP — abono inicial que se devuelve
    MONEDA          = "COP"

    @staticmethod
    def procesar_reembolso(id_transaccion: str,
                           monto: float = 5000.0,
                           moneda: str = "COP") -> bool:
        logging.info(
            f"[PASARELA DE PAGO] Procesando reembolso de ${monto:,.0f} {moneda} "
            f"para TRX: {id_transaccion}..."
        )
        return True   # Simulación: siempre exitoso

    @staticmethod
    def crear_sesion_checkout(reserva_id: str, monto: float, moneda: str = "COP") -> str:
        """Genera el link de checkout para el abono (integrar con Wompi/MercadoPago)."""
        logging.info(
            f"[PASARELA] Creando checkout para reserva {reserva_id} "
            f"por ${monto:,.0f} {moneda}..."
        )
        # TODO: Reemplazar por URL real de la pasarela
        return f"https://checkout.pasarela.com/pay/{reserva_id}?amount={int(monto)}"

    @staticmethod
    def validar_transaccion(monto_requerido: float, monto_recibido: float) -> bool:
        """Valida que el monto recibido cubre el abono requerido."""
        return monto_recibido >= monto_requerido


# ============================================================================
# ESTADO DE RESERVA — Máquina de estados del flujo de agendamiento
# ============================================================================

class EstadoReserva(Enum):
    INICIADA          = "Iniciada"
    DATOS_REGISTRADOS = "Datos Registrados"
    PENDIENTE_PAGO    = "Pendiente de Pago"
    CONFIRMADA        = "Confirmada"
    ERROR_PAGO        = "Error en Pago"


class ReservaError(Exception):
    """Excepción base para errores en el proceso de reserva."""
    pass


@dataclass(frozen=True)
class ClienteDTO:
    """Data Transfer Object para validar y transportar los datos del cliente."""
    nombre:   str
    telefono: str
    correo:   str

    def __post_init__(self):
        if not self.nombre.strip():
            raise ValueError("El nombre no puede estar vacío.")
        if "@" not in self.correo:
            raise ValueError("El correo electrónico no es válido.")


class ReservaService:
    """
    Orquesta el flujo completo de reserva en 6 pasos:
    1. Cliente inicia la reserva
    2. Captura de datos del cliente
    3. Redirección a pasarela de pago
    4. Procesamiento del pago
    5. Validación y confirmación
    6. Generación del comprobante descargable
    """
    ABONO_REQUERIDO: float = 5000.0  # $5.000 COP

    def __init__(self, servicio: str, fecha_cita: datetime):
        self.id_reserva:      str                  = f"RES-{uuid.uuid4().hex[:6].upper()}"
        self.servicio:        str                  = servicio
        self.fecha_cita:      datetime             = fecha_cita
        self.estado:          EstadoReserva        = EstadoReserva.INICIADA
        self.cliente:         Optional[ClienteDTO] = None
        self.id_transaccion:  Optional[str]        = None
        logging.info(
            f"Paso 1: Cliente inició la reserva [{self.id_reserva}] "
            f"para '{self.servicio}'."
        )

    def registrar_datos_cliente(self, nombre: str, telefono: str, correo: str) -> str:
        """Pasos 2 y 3: Valida los datos del cliente y genera link de pago."""
        if self.estado != EstadoReserva.INICIADA:
            raise ReservaError("No se pueden actualizar datos en el estado actual.")

        self.cliente = ClienteDTO(nombre=nombre, telefono=telefono, correo=correo)
        self.estado  = EstadoReserva.DATOS_REGISTRADOS
        logging.info(f"Paso 2: Datos de '{self.cliente.nombre}' capturados correctamente.")

        self.estado  = EstadoReserva.PENDIENTE_PAGO
        link_pago    = PasarelaPagoService.crear_sesion_checkout(
            self.id_reserva, self.ABONO_REQUERIDO
        )
        logging.info("Paso 3: Redirigiendo cliente a la pasarela de pagos...")
        return link_pago

    def recibir_confirmacion_pago(self, monto_pagado: float, transaction_id: str) -> bool:
        """Pasos 4, 5 y 6: Procesa y valida el pago, cambia el estado."""
        if self.estado != EstadoReserva.PENDIENTE_PAGO:
            raise ReservaError("La reserva no está esperando un pago.")

        logging.info(
            f"Paso 4: Procesando pago de ${monto_pagado:,.0f} COP "
            f"(Ref: {transaction_id})..."
        )
        pago_exitoso = PasarelaPagoService.validar_transaccion(
            self.ABONO_REQUERIDO, monto_pagado
        )
        if pago_exitoso:
            self.id_transaccion = transaction_id
            self.estado         = EstadoReserva.CONFIRMADA
            logging.info(
                f"Paso 5 y 6: Pago confirmado. Estado → '{self.estado.value}'."
            )
            return True
        else:
            self.estado = EstadoReserva.ERROR_PAGO
            logging.error("Paso 5: Pago rechazado o monto inferior al requerido.")
            return False

    def obtener_comprobante_cita(self) -> Dict[str, Any]:
        """Genera el diccionario de datos para el comprobante PDF."""
        if self.estado != EstadoReserva.CONFIRMADA:
            raise ReservaError("Solo se puede generar comprobante de citas confirmadas.")
        return {
            "comprobante_id": f"DOC-{self.id_reserva}",
            "estado":         self.estado.value,
            "detalles_servicio": {
                "servicio":   self.servicio,
                "fecha_hora": self.fecha_cita.strftime("%Y-%m-%d %H:%M"),
            },
            "cliente": {
                "nombre":   self.cliente.nombre,
                "telefono": self.cliente.telefono,
                "correo":   self.cliente.correo,
            },
            "pago": {
                "monto_abono":    f"${self.ABONO_REQUERIDO:,.0f} COP",
                "id_transaccion": self.id_transaccion,
            }
        }


# ============================================================================
# AGENDA DIARIA — SistemaAgendaDiaria, CitaDiaria, EstadoCitaOperativa
# ============================================================================

class EstadoCitaOperativa(Enum):
    PROGRAMADA  = "Programada"
    EN_ATENCION = "En atención"
    COMPLETADA  = "Completada"
    CANCELADA   = "Cancelada"


class MetodoPagoSaldo(Enum):
    EFECTIVO      = "Efectivo"
    TRANSFERENCIA = "Transferencia"
    NEQUI         = "Nequi"
    DAVIPLATA     = "Daviplata"
    TARJETA       = "Tarjeta"


class InvalidOperationError(Exception):
    """Error cuando se intenta una operación inválida sobre el estado de una cita."""
    pass


@dataclass
class CitaDiaria:
    """Representa una cita operativa del día con su ciclo de vida completo."""
    id_cita:            str
    cliente_nombre:     str
    profesional_id:     str
    profesional_nombre: str
    servicio_nombre:    str
    hora:               time
    precio_total:       float
    abono_previo:       float                    = 5000.0
    estado:             EstadoCitaOperativa      = field(default_factory=lambda: EstadoCitaOperativa.PROGRAMADA)
    metodo_pago_saldo:  Optional[MetodoPagoSaldo]= None
    saldo_pagado:       float                    = 0.0

    @property
    def saldo_pendiente(self) -> float:
        """Paso 4: Calcula saldo pendiente = Precio Total - Abono."""
        return max(0.0, self.precio_total - self.abono_previo)


class SistemaAgendaDiaria:
    """
    Módulo de gestión de agenda diaria y liquidación de citas.
    Pasos:
    1. Acceso a la vista Agenda Diaria
    2. Visualización en cuadrícula filtrada por profesional
    3. Marcar cita como 'En atención' cuando llega el cliente
    4. Calcular saldo pendiente al finalizar el servicio
    5. Registrar pago del saldo y marcar como 'Completada'
    """

    def __init__(self):
        self._citas_db: List[CitaDiaria] = []

    def cargar_desde_bd(self, citas_orm: list) -> None:
        """Carga citas desde los objetos ORM de SQLAlchemy."""
        self._citas_db = []
        for c in citas_orm:
            empleado  = c.empleado
            servicio  = c.servicio
            cliente   = c.cliente
            self._citas_db.append(CitaDiaria(
                id_cita            = str(c.id_cita),
                cliente_nombre     = cliente.nombre   if cliente  else 'N/A',
                profesional_id     = str(empleado.id_empleado) if empleado else '0',
                profesional_nombre = empleado.nombre  if empleado else 'Sin asignar',
                servicio_nombre    = servicio.nombre_servicio if servicio else 'N/A',
                hora               = c.fecha_hora_inicio.time(),
                precio_total       = float(c.monto_total or 0),
                abono_previo       = float(c.monto_abono or 5000),
                estado             = self._mapear_estado(c.estado),
                saldo_pagado       = float(c.monto_total or 0) - float(c.saldo_pendiente or 0)
                                     if c.estado == 'completada' else 0.0,
            ))

    @staticmethod
    def _mapear_estado(estado_bd: str) -> EstadoCitaOperativa:
        return {
            'pendiente_pago': EstadoCitaOperativa.PROGRAMADA,
            'confirmada':     EstadoCitaOperativa.PROGRAMADA,
            'en_atencion':    EstadoCitaOperativa.EN_ATENCION,
            'completada':     EstadoCitaOperativa.COMPLETADA,
            'cancelada':      EstadoCitaOperativa.CANCELADA,
            'no_asistio':     EstadoCitaOperativa.CANCELADA,
        }.get(estado_bd, EstadoCitaOperativa.PROGRAMADA)

    def obtener_cuadrilla_agenda_diaria(self, fecha_consulta: date) -> Dict[str, List[Dict]]:
        """Pasos 1 y 2: Genera la cuadrícula diaria agrupada por profesional."""
        logging.info(
            f"[AGENDA DIARIA] Generando cuadrícula para el {fecha_consulta.strftime('%d/%m/%Y')}..."
        )
        cuadricula: Dict[str, List[Dict]] = {}
        for cita in self._citas_db:
            if cita.profesional_nombre not in cuadricula:
                cuadricula[cita.profesional_nombre] = []
            cuadricula[cita.profesional_nombre].append({
                "id_cita":         cita.id_cita,
                "hora":            cita.hora.strftime("%H:%M"),
                "cliente":         cita.cliente_nombre,
                "servicio":        cita.servicio_nombre,
                "estado":          cita.estado.value,
                "estado_key":      cita.estado.name,
                "precio_total":    f"${cita.precio_total:,.0f}",
                "abono":           f"${cita.abono_previo:,.0f}",
                "saldo_pendiente": f"${cita.saldo_pendiente:,.0f}",
                "saldo_num":       cita.saldo_pendiente,
            })
        return cuadricula

    def marcar_en_atencion(self, id_cita: str) -> bool:
        """Paso 3: El cliente llegó — cambia estado a 'En atención'."""
        cita = self._buscar_cita(id_cita)
        if not cita:
            return False
        if cita.estado != EstadoCitaOperativa.PROGRAMADA:
            logging.warning(f"[AGENDA] Cita {id_cita} no está PROGRAMADA.")
            return False
        cita.estado = EstadoCitaOperativa.EN_ATENCION
        logging.info(
            f"[AGENDA] Paso 3: {cita.cliente_nombre} llegó. "
            f"Cita {id_cita} → 'En atención'."
        )
        return True

    def liquidar_y_completar_cita(self,
                                   id_cita: str,
                                   metodo_pago: MetodoPagoSaldo) -> Dict[str, float]:
        """Pasos 4 y 5: Calcula saldo, registra pago y marca como 'Completada'."""
        cita = self._buscar_cita(id_cita)
        if not cita:
            raise ValueError(f"Cita {id_cita} no encontrada.")
        if cita.estado != EstadoCitaOperativa.EN_ATENCION:
            raise InvalidOperationError(
                "La cita debe estar 'En atención' para poder ser completada."
            )
        monto_a_cobrar     = cita.saldo_pendiente
        cita.saldo_pagado  = monto_a_cobrar
        cita.metodo_pago_saldo = metodo_pago
        cita.estado        = EstadoCitaOperativa.COMPLETADA
        logging.info(
            f"[AGENDA] Pasos 4-5: Saldo ${monto_a_cobrar:,.0f} COP cobrado "
            f"en [{metodo_pago.value}]. Cita {id_cita} → 'Completada'."
        )
        return {
            "precio_total":  cita.precio_total,
            "abono_previo":  cita.abono_previo,
            "saldo_cobrado": monto_a_cobrar,
            "metodo_pago":   metodo_pago.value
        }

    def _buscar_cita(self, id_cita: str) -> Optional[CitaDiaria]:
        for c in self._citas_db:
            if c.id_cita == id_cita:
                return c
        logging.error(f"[AGENDA] Cita {id_cita} no encontrada.")
        return None


# ============================================================================
# GESTIÓN Y REPROGRAMACIÓN DE CITAS — SistemaGestionCitas
# ============================================================================

class ReprogramacionError(Exception):
    """Excepción para errores en el flujo de reprogramación."""
    pass


class ServicioNotificaciones:
    """
    Simula el envío de actualizaciones al cliente
    vía Email y WhatsApp (integrar con API real en producción).
    """

    @staticmethod
    def enviar_confirmacion_reprogramacion(
        nombre:              str,
        telefono:            str,
        email:               str,
        nueva_fecha:         datetime,
        profesional_nombre:  str,
        abono:               float
    ) -> None:
        mensaje = (
            f"Hola {nombre}, tu cita ha sido reprogramada con éxito.\n"
            f"Nueva Fecha/Hora: {nueva_fecha.strftime('%d/%m/%Y a las %H:%M')}\n"
            f"Profesional: {profesional_nombre}\n"
            f"Se conserva tu abono de: ${abono:,.0f} COP."
        )
        # Log simulando WhatsApp — reemplazar por API de WhatsApp Business en producción
        logging.info(
            f"[NOTIFICACIÓN WHATSAPP] → {telefono}: {mensaje}"
        )
        logging.info(
            f"[NOTIFICACIÓN EMAIL]    → {email}: {mensaje}"
        )


class SistemaGestionCitas:
    """
    Módulo central para gestión y reprogramación de citas.
    Flujo de 6 pasos:
    1. Cliente accede a 'Gestionar Cita' desde el link de su token
    2. Selecciona la opción 'Reprogramar Cita'
    3. Sistema evalúa política de tiempo (>= 2 horas de anticipación)
    4. Despliega calendario con disponibilidad del profesional
    5. Cliente selecciona nueva fecha/hora (y opcionalmente nuevo profesional)
    6. Sistema actualiza la cita, mantiene abono y notifica al cliente
    """
    HORAS_MINIMAS_REPROGRAMACION = 2
    ABONO_COP                    = 5000.0

    @staticmethod
    def acceder_a_gestion(token: str, cita_orm) -> dict:
        """Paso 1: Valida el token y retorna los datos de la cita."""
        if not cita_orm:
            raise ReprogramacionError("Enlace de gestión inválido o expirado.")
        logging.info(
            f"[REPROGRAMACIÓN] Paso 1: Cliente "
            f"'{cita_orm.cliente.nombre if cita_orm.cliente else 'N/A'}' "
            f"accedió a Gestionar Cita (token: {token[:8]}...)."
        )
        return {
            "id_cita":       cita_orm.id_cita,
            "estado":        cita_orm.estado,
            "fecha_hora":    cita_orm.fecha_hora_inicio,
            "profesional":   cita_orm.empleado.nombre if cita_orm.empleado else 'Sin asignar',
            "servicio":      cita_orm.servicio.nombre_servicio if cita_orm.servicio else 'N/A',
            "abono":         float(cita_orm.monto_abono or SistemaGestionCitas.ABONO_COP),
            "token":         token,
        }

    @staticmethod
    def validar_politica_reprogramacion(cita_orm,
                                        hora_actual: Optional[datetime] = None) -> bool:
        """Pasos 2 y 3: Valida que falten >= 2 horas para la cita."""
        logging.info(
            f"[REPROGRAMACIÓN] Paso 2: Cliente seleccionó 'Reprogramar Cita'."
        )
        if hora_actual is None:
            hora_actual = datetime.now()

        tiempo_restante = cita_orm.fecha_hora_inicio - hora_actual
        horas_restantes = tiempo_restante.total_seconds() / 3600

        if horas_restantes < SistemaGestionCitas.HORAS_MINIMAS_REPROGRAMACION:
            logging.error(
                f"[REPROGRAMACIÓN] Paso 3 [DENEGADO]: {horas_restantes:.2f} hrs restantes. "
                f"Se requieren >= {SistemaGestionCitas.HORAS_MINIMAS_REPROGRAMACION} hrs."
            )
            raise ReprogramacionError(
                f"No es posible reprogramar con menos de "
                f"{SistemaGestionCitas.HORAS_MINIMAS_REPROGRAMACION} horas de anticipación. "
                f"Tiempo restante: {horas_restantes:.1f} hrs."
            )

        logging.info(
            f"[REPROGRAMACIÓN] Paso 3 [APROBADO]: "
            f"{horas_restantes:.2f} hrs restantes — cumple política de tiempo."
        )
        return True

    @staticmethod
    def obtener_agenda_disponible(id_servicio: int,
                                  id_empleado_actual: Optional[int],
                                  fecha_desde: Optional[datetime] = None) -> dict:
        """
        Paso 4: Construye el mapa de disponibilidad de empleados
        para el servicio dado a partir de hoy + 2 horas.
        Retorna {empleado_nombre: [slots de datetime disponibles]}
        """
        from sqlalchemy import func as sqlfunc
        if fecha_desde is None:
            fecha_desde = datetime.now() + timedelta(
                hours=SistemaGestionCitas.HORAS_MINIMAS_REPROGRAMACION
            )

        logging.info(
            f"[REPROGRAMACIÓN] Paso 4: Desplegando disponibilidad "
            f"desde {fecha_desde.strftime('%d/%m/%Y %H:%M')}..."
        )
        # Obtener empleados que realizan este servicio
        from app import db, EmpleadoServicio, Empleado, HorarioEmpleado, Cita as CitaModel
        empleados_ids = db.session.query(EmpleadoServicio.id_empleado)\
            .filter_by(id_servicio=id_servicio).all()
        empleados_ids = [e[0] for e in empleados_ids]

        disponibilidad = {}
        for emp_id in empleados_ids:
            empleado = Empleado.query.get(emp_id)
            if not empleado or not empleado.activo:
                continue
            # Horarios de los próximos 7 días
            slots = []
            for dias_offset in range(0, 8):
                dia = (fecha_desde + timedelta(days=dias_offset)).date()
                dia_semana = (dia.weekday() + 1) % 7  # 0=Domingo
                horario = HorarioEmpleado.query.filter_by(
                    id_empleado=emp_id, dia_semana=dia_semana
                ).first()
                if not horario:
                    continue
                slot = datetime.combine(dia, horario.hora_inicio)
                fin  = datetime.combine(dia, horario.hora_fin)
                while slot < fin:
                    if slot >= fecha_desde:
                        # Verificar si el slot está libre
                        ocupado = CitaModel.query.filter(
                            CitaModel.id_empleado == emp_id,
                            CitaModel.fecha_hora_inicio <= slot,
                            CitaModel.fecha_hora_fin    >  slot,
                            CitaModel.estado.in_(['pendiente_pago','confirmada','en_atencion'])
                        ).first()
                        if not ocupado:
                            slots.append(slot)
                    slot += timedelta(minutes=30)
            if slots:
                disponibilidad[emp_id] = {
                    "nombre": empleado.nombre,
                    "actual": emp_id == id_empleado_actual,
                    "slots":  [s.strftime('%Y-%m-%d %H:%M') for s in slots[:20]]
                }
        return disponibilidad

    @staticmethod
    def ejecutar_reprogramacion(cita_orm,
                                nueva_fecha_hora: datetime,
                                nuevo_id_empleado: Optional[int] = None) -> dict:
        """
        Pasos 5 y 6: Actualiza la cita en BD,
        mantiene el abono y notifica al cliente.
        """
        profesional_id  = nuevo_id_empleado or cita_orm.id_empleado
        empleado        = Empleado.query.get(profesional_id) if profesional_id else None
        fecha_fin       = nueva_fecha_hora + timedelta(
            minutes=cita_orm.servicio.duracion_minutos if cita_orm.servicio else 60
        )

        # Actualizar cita
        cita_orm.fecha_hora_inicio = nueva_fecha_hora
        cita_orm.fecha_hora_fin    = fecha_fin
        cita_orm.id_empleado       = profesional_id
        cita_orm.estado            = 'confirmada'   # Reprogramada → vuelve a confirmada

        from app import db
        db.session.commit()

        profesional_nombre = empleado.nombre if empleado else 'Sin asignar'
        abono              = float(cita_orm.monto_abono or SistemaGestionCitas.ABONO_COP)

        logging.info(
            f"[REPROGRAMACIÓN] Paso 6: Cita #{cita_orm.id_cita} reprogramada → "
            f"{nueva_fecha_hora.strftime('%d/%m/%Y %H:%M')} "
            f"con {profesional_nombre}. Abono ${abono:,.0f} COP mantenido."
        )

        # Notificar al cliente (Email + WhatsApp)
        if cita_orm.cliente:
            ServicioNotificaciones.enviar_confirmacion_reprogramacion(
                nombre             = cita_orm.cliente.nombre,
                telefono           = cita_orm.cliente.telefono,
                email              = cita_orm.cliente.email,
                nueva_fecha        = nueva_fecha_hora,
                profesional_nombre = profesional_nombre,
                abono              = abono
            )

        return {
            "id_cita":            cita_orm.id_cita,
            "nueva_fecha":        nueva_fecha_hora.strftime('%d/%m/%Y a las %H:%M'),
            "profesional":        profesional_nombre,
            "abono_mantenido":    abono,
            "codigo_reserva":     cita_orm.codigo_reserva,
        }


# ============================================================================
# MODELOS DE LA BASE DE DATOS CON RELACIONES
# ============================================================================


class Usuario(db.Model):
    """Usuarios del sistema (clientes y administradores)"""
    __tablename__ = 'usuario'

    id = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), nullable=False)
    email = db.Column(db.String(120), unique=True, nullable=False)
    telefono = db.Column(db.String(20), nullable=False)
    password = db.Column(db.String(200), nullable=False)
    tipo_usuario = db.Column(db.String(20), nullable=False)
    fecha_registro = db.Column(db.DateTime, default=datetime.utcnow)
    activo = db.Column(db.Boolean, default=True)

    # Relaciones
    citas = db.relationship('Cita', backref='cliente', lazy=True, foreign_keys='Cita.id_cliente')
    # Notificaciones del usuario
    notificaciones = db.relationship('Notificacion', backref='usuario', lazy=True)

    def __repr__(self):
        return f'<Usuario {self.nombre} - {self.tipo_usuario}>'


class Servicio(db.Model):
    """Servicios ofrecidos por el salón"""
    __tablename__ = 'servicios'

    id_servicio = db.Column(db.Integer, primary_key=True)
    nombre_servicio = db.Column(db.String(100), nullable=False)
    descripcion = db.Column(db.Text)
    precio_total = db.Column(db.Numeric(10, 2), nullable=False)
    duracion_minutos = db.Column(db.Integer, nullable=False)
    activo = db.Column(db.Boolean, default=True)

    # Relaciones
    citas = db.relationship('Cita', backref='servicio', lazy=True)
    empleados = db.relationship('Empleado', secondary='empleado_servicios', backref='servicios')

    def __repr__(self):
        return f'<Servicio {self.nombre_servicio}>'


class Empleado(db.Model):
    """Empleados del salón"""
    __tablename__ = 'empleados'

    id_empleado = db.Column(db.Integer, primary_key=True)
    nombre = db.Column(db.String(100), nullable=False)
    especialidad = db.Column(db.String(100))
    activo = db.Column(db.Boolean, default=True)
    fecha_registro = db.Column(db.DateTime, default=datetime.utcnow)

    # Relaciones
    horarios = db.relationship('HorarioEmpleado', backref='empleado', lazy=True, cascade='all, delete-orphan')
    citas = db.relationship('Cita', backref='empleado', lazy=True)

    def __repr__(self):
        return f'<Empleado {self.nombre}>'


class EmpleadoServicio(db.Model):
    """Relación empleados-servicios (Tabla intermedia Many-to-Many)"""
    __tablename__ = 'empleado_servicios'

    id_empleado = db.Column(db.Integer, db.ForeignKey('empleados.id_empleado', ondelete='CASCADE'), primary_key=True)
    id_servicio = db.Column(db.Integer, db.ForeignKey('servicios.id_servicio', ondelete='CASCADE'), primary_key=True)


class HorarioEmpleado(db.Model):
    """Horarios de trabajo de los empleados"""
    __tablename__ = 'horarios_empleados'

    id_horario = db.Column(db.Integer, primary_key=True)
    id_empleado = db.Column(db.Integer, db.ForeignKey('empleados.id_empleado', ondelete='CASCADE'), nullable=False)
    dia_semana = db.Column(db.Integer, nullable=False)  # 0=Domingo, 1=Lunes, ..., 6=Sábado
    hora_inicio = db.Column(db.Time, nullable=False)
    hora_fin = db.Column(db.Time, nullable=False)

    def __repr__(self):
        return f'<HorarioEmpleado {self.empleado.nombre if self.empleado else "N/A"} - Día {self.dia_semana}>'


class Cita(db.Model):
    """Citas agendadas"""
    __tablename__ = 'citas'

    id_cita = db.Column(db.Integer, primary_key=True)
    id_cliente = db.Column(db.Integer, db.ForeignKey('usuario.id', ondelete='CASCADE'), nullable=False)
    id_empleado = db.Column(db.Integer, db.ForeignKey('empleados.id_empleado', ondelete='SET NULL'))
    id_servicio = db.Column(db.Integer, db.ForeignKey('servicios.id_servicio', ondelete='RESTRICT'), nullable=False)
    fecha_hora_inicio = db.Column(db.DateTime, nullable=False)
    fecha_hora_fin = db.Column(db.DateTime, nullable=False)
    monto_total = db.Column(db.Numeric(10, 2))
    monto_abono = db.Column(db.Numeric(10, 2))
    saldo_pendiente = db.Column(db.Numeric(10, 2))
    estado = db.Column(
        db.Enum(
            'pendiente_pago',
            'confirmada',
            'en_atencion',
            'completada',
            'cancelada',
            'no_asistio',
            name='estado_cita_enum',
            native_enum=True,
            create_constraint=False,
        ),
        nullable=False,
        default='pendiente_pago',
    )
    reembolsado    = db.Column(db.Boolean, default=False)
    codigo_reserva = db.Column(db.String(10), unique=True)
    token_gestion  = db.Column(db.String(32), unique=True, nullable=True)  # Token para link de gestión
    notas          = db.Column(db.Text)
    fecha_creacion = db.Column(db.DateTime, default=datetime.utcnow)

    # Relación con pagos (una cita tiene máximo un pago)
    pago = db.relationship('Pago', backref='cita', uselist=False, lazy=True)

    def __repr__(self):
        return f'<Cita {self.id_cita} - {self.estado}>'


class Pago(db.Model):
    """Pagos registrados por cada cita"""
    __tablename__ = 'pagos'

    id_pago = db.Column(db.Integer, primary_key=True)
    id_cita = db.Column(db.Integer, db.ForeignKey('citas.id_cita', ondelete='CASCADE'),
                        nullable=False, unique=True)
    monto = db.Column(db.Numeric(10, 2), nullable=False)
    metodo_pago = db.Column(
        db.Enum(
            'efectivo',
            'tarjeta',
            'transferencia',
            'nequi',
            'daviplata',
            name='metodo_pago_enum',
            native_enum=True,
            create_constraint=False,
        ),
        nullable=False,
        default='efectivo',
    )
    estado_pago = db.Column(db.String(20), nullable=False, default='completado')
    referencia = db.Column(db.String(100))
    fecha_pago = db.Column(db.DateTime, default=datetime.utcnow)
    notas = db.Column(db.Text)

    def __repr__(self):
        return f'<Pago {self.id_pago} - Cita {self.id_cita} - ${self.monto}>'


class Notificacion(db.Model):
    """Notificaciones para usuarios"""
    __tablename__ = 'notificaciones'

    id = db.Column(db.Integer, primary_key=True)
    id_usuario = db.Column(db.Integer, db.ForeignKey('usuario.id', ondelete='CASCADE'), nullable=False)
    titulo = db.Column(db.String(200), nullable=False)
    mensaje = db.Column(db.Text)
    target = db.Column(db.String(300))
    leido = db.Column(db.Boolean, default=False)
    fecha = db.Column(db.DateTime, default=datetime.utcnow)

    def __repr__(self):
        return f'<Notificacion {self.id} -> Usuario {self.id_usuario} - {self.titulo}>'


# Crear las tablas
with app.app_context():
    db.create_all()
    # Crear usuario administrador por defecto si no existe
    _admin_email = os.environ.get('ADMIN_EMAIL', 'admin@rossmix.com')
    _admin_pwd   = os.environ.get('ADMIN_PASSWORD')
    admin = Usuario.query.filter_by(email=_admin_email).first()
    if not admin:
        if not _admin_pwd:
            print('Aviso: no se creó admin por defecto — define ADMIN_PASSWORD en .env')
        else:
            admin = Usuario(
                nombre='Administrador',
                email=_admin_email,
                telefono='3000000000',
                password=generate_password_hash(_admin_pwd),
                tipo_usuario='admin'
            )
            db.session.add(admin)
            db.session.commit()
            print(f'Usuario administrador creado: {_admin_email}')

# ============================================================================
# DECORADOR PARA RUTAS DE ADMINISTRADOR
# ============================================================================


def admin_required(f):
    """Decorador para requerir acceso de administrador"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'usuario_id' not in session:
            flash('Debes iniciar sesión', 'error')
            return redirect(url_for('login'))
        if session.get('tipo_usuario') != 'admin':
            flash('No tienes permisos para acceder a esta sección', 'error')
            return redirect(url_for('dashboard_cliente'))
        return f(*args, **kwargs)
    return decorated_function

# Rutas


@app.route('/')
def index():
    return render_template('index.html')

    # Helper: crear notificación


def add_notificacion(id_usuario, titulo, mensaje=None, target=None):
    try:
        n = Notificacion(id_usuario=id_usuario, titulo=titulo, mensaje=mensaje, target=target)
        db.session.add(n)
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        print('Error al crear notificacion:', e)

    # Inyectar notificaciones en templates


@app.context_processor
def inject_notificaciones():
    if 'usuario_id' in session:
        try:
            notifs = Notificacion.query.filter_by(
                id_usuario=session['usuario_id']).order_by(
                Notificacion.fecha.desc()).limit(6).all()
            unread = Notificacion.query.filter_by(id_usuario=session['usuario_id'], leido=False).count()
            return dict(notificaciones=notifs, notificaciones_unread=unread)
        except Exception:
            return dict(notificaciones=[], notificaciones_unread=0)
    return dict(notificaciones=[], notificaciones_unread=0)


@app.route('/notificaciones')
def notificaciones():
    if 'usuario_id' not in session:
        flash('Debes iniciar sesión', 'error')
        return redirect(url_for('login'))
    # Paginación simple
    page = request.args.get('page', 1, type=int)
    per_page = 20
    q = Notificacion.query.filter_by(id_usuario=session['usuario_id']).order_by(Notificacion.fecha.desc())
    total = q.count()
    total_pages = (total + per_page - 1) // per_page
    notifs = q.offset((page - 1) * per_page).limit(per_page).all()
    return render_template('notificaciones.html', notificaciones=notifs, page=page, total_pages=total_pages)


@app.route('/test-image')
def test_image():
    return render_template('test_image.html')


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form.get('email')
        password = request.form.get('password')

        usuario = Usuario.query.filter_by(email=email).first()

        if usuario and check_password_hash(usuario.password, password):
            if not usuario.activo:
                flash('Tu cuenta está desactivada. Contacta al administrador.', 'error')
                return redirect(url_for('login'))

            session['usuario_id'] = usuario.id
            session['nombre'] = usuario.nombre
            session['email'] = usuario.email
            session['tipo_usuario'] = usuario.tipo_usuario
            flash(f'¡Bienvenido/a {usuario.nombre}!', 'success')

            if usuario.tipo_usuario == 'admin':
                return redirect(url_for('dashboard_admin'))
            else:
                return redirect(url_for('dashboard_cliente'))
        else:
            flash('Email o contraseña incorrectos', 'error')

    return render_template('login.html')


@app.route('/registro', methods=['GET', 'POST'])
def registro():
    if request.method == 'POST':
        nombre = request.form.get('nombre')
        email = request.form.get('email')
        telefono = request.form.get('telefono')
        password = request.form.get('password')
        confirmar_password = request.form.get('confirmar_password')

        # Validaciones
        if not all([nombre, email, telefono, password, confirmar_password]):
            flash('Todos los campos son obligatorios', 'error')
            return redirect(url_for('registro'))

        if password != confirmar_password:
            flash('Las contraseñas no coinciden', 'error')
            return redirect(url_for('registro'))

        if len(password) < 6:
            flash('La contraseña debe tener al menos 6 caracteres', 'error')
            return redirect(url_for('registro'))

        # Verificar si el email ya existe
        usuario_existente = Usuario.query.filter_by(email=email).first()
        if usuario_existente:
            flash('Este email ya está registrado', 'error')
            return redirect(url_for('registro'))

        # Crear nuevo usuario
        nuevo_usuario = Usuario(
            nombre=nombre,
            email=email,
            telefono=telefono,
            password=generate_password_hash(password),
            tipo_usuario='cliente'
        )

        db.session.add(nuevo_usuario)
        db.session.commit()

        # ── Log en consola confirmando escritura en PostgreSQL ──────────────
        logging.info(
            f"[NUEVO USUARIO] ID: {nuevo_usuario.id} | "
            f"Nombre: {nuevo_usuario.nombre} | "
            f"Email: {nuevo_usuario.email} | "
            f"Teléfono: {nuevo_usuario.telefono} | "
            f"Fecha: {nuevo_usuario.fecha_registro.strftime('%d/%m/%Y %H:%M:%S')} | "
            f"BD: PostgreSQL — INSERT exitoso ✓"
        )

        # ── Notificar a todos los administradores en tiempo real ────────────
        try:
            admins = Usuario.query.filter_by(tipo_usuario='admin').all()
            for admin in admins:
                add_notificacion(
                    admin.id,
                    f'🆕 Nueva clienta registrada',
                    f'{nuevo_usuario.nombre} ({nuevo_usuario.email} | {nuevo_usuario.telefono}) '
                    f'se registró el {nuevo_usuario.fecha_registro.strftime("%d/%m/%Y a las %H:%M")}.',
                    target=url_for('admin_clientes')
                )
        except Exception as e:
            logging.error(f"[NUEVO USUARIO] Error al notificar admins: {e}")

        flash('¡Registro exitoso! Ya puedes iniciar sesión', 'success')
        return redirect(url_for('login'))

    return render_template('registro.html')


@app.route('/dashboard/admin')
@admin_required
def dashboard_admin():
    """Dashboard principal de administrador con estadísticas"""
    from sqlalchemy import func

    # Citas de hoy
    hoy = datetime.now().date()
    citas_hoy = Cita.query.filter(
        func.date(Cita.fecha_hora_inicio) == hoy,
        Cita.estado.in_(['pendiente_pago', 'confirmada', 'en_atencion'])
    ).count()

    # Total clientes
    total_clientes = Usuario.query.filter_by(tipo_usuario='cliente', activo=True).count()

    # Empleados activos
    empleados_activos = Empleado.query.filter_by(activo=True).count()

    # Ingresos del mes
    primer_dia_mes = datetime.now().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    ingresos_mes = db.session.query(func.sum(Cita.monto_total)).filter(
        Cita.fecha_creacion >= primer_dia_mes,
        Cita.estado.in_(['completada', 'confirmada'])
    ).scalar() or 0

    stats = {
        'citas_hoy': citas_hoy,
        'total_clientes': total_clientes,
        'empleados_activos': empleados_activos,
        'ingresos_mes': ingresos_mes,
        'pagos_pendientes': Cita.query.filter_by(estado='pendiente_pago').count()
    }

    return render_template('dashboard_admin.html', stats=stats)


@app.route('/dashboard/cliente')
def dashboard_cliente():
    if 'usuario_id' not in session or session.get('tipo_usuario') != 'cliente':
        flash('Debes iniciar sesión como cliente', 'error')
        return redirect(url_for('login'))

    id_cliente = session['usuario_id']

    # Citas pendientes/confirmadas (futuras)
    citas_pendientes = Cita.query.filter(
        Cita.id_cliente == id_cliente,
        Cita.fecha_hora_inicio >= datetime.now(),
        Cita.estado.in_(['pendiente_pago', 'confirmada'])
    ).count()

    # Citas completadas
    citas_completadas = Cita.query.filter(
        Cita.id_cliente == id_cliente,
        Cita.estado == 'completada'
    ).count()

    # Próxima cita
    proxima_cita = db.session.query(Cita, Servicio, Empleado).join(
        Servicio, Cita.id_servicio == Servicio.id_servicio
    ).outerjoin(
        Empleado, Cita.id_empleado == Empleado.id_empleado
    ).filter(
        Cita.id_cliente == id_cliente,
        Cita.fecha_hora_inicio >= datetime.now(),
        Cita.estado.in_(['pendiente_pago', 'confirmada'])
    ).order_by(Cita.fecha_hora_inicio).first()

    stats = {
        'citas_pendientes': citas_pendientes,
        'citas_completadas': citas_completadas
    }

    return render_template('dashboard_cliente.html', stats=stats, proxima_cita=proxima_cita)


@app.route('/logout')
def logout():
    session.clear()
    flash('Has cerrado sesión correctamente', 'success')
    return redirect(url_for('index'))


# Marcar una notificación como leída
@app.route('/notificaciones/marcar-leida/<int:notif_id>', methods=['POST'])
def notificacion_marcar_leida(notif_id):
    if 'usuario_id' not in session:
        return jsonify({'success': False, 'message': 'No autorizado'}), 401
    n = Notificacion.query.get_or_404(notif_id)
    # permitir solo al propietario de la notificación o a admins
    if n.id_usuario != session['usuario_id'] and session.get('tipo_usuario') != 'admin':
        return jsonify({'success': False, 'message': 'No autorizado'}), 403
    n.leido = True
    db.session.commit()
    # devolver nuevo conteo de no leídos
    unread = Notificacion.query.filter_by(id_usuario=session['usuario_id'], leido=False).count()
    return jsonify({'success': True, 'unread': unread})


# Marcar todas las notificaciones del usuario como leídas
@app.route('/notificaciones/marcar-todas', methods=['POST'])
def notificaciones_marcar_todas():
    if 'usuario_id' not in session:
        return jsonify({'success': False, 'message': 'No autorizado'}), 401
    try:
        Notificacion.query.filter_by(id_usuario=session['usuario_id'], leido=False).update({'leido': True})
        db.session.commit()
    except Exception as e:
        db.session.rollback()
        return jsonify({'success': False, 'message': str(e)}), 500
    return jsonify({'success': True, 'unread': 0})

# ============================================================================
# RUTAS DEL SISTEMA DE CITAS
# ============================================================================


@app.route('/citas/agendar/paso1')
def agendar_paso1():
    """Paso 1: Seleccionar servicio"""
    if 'usuario_id' not in session:
        flash('Debes iniciar sesión para agendar una cita', 'error')
        return redirect(url_for('login'))

    # Obtener todos los servicios activos
    servicios = Servicio.query.filter_by(activo=True).order_by(Servicio.nombre_servicio).all()

    return render_template('citas/paso1_servicio.html', servicios=servicios)


@app.route('/citas/agendar/paso2/<int:id_servicio>')
def agendar_paso2(id_servicio):
    """Paso 2: Seleccionar empleado o aleatorio"""
    if 'usuario_id' not in session:
        flash('Debes iniciar sesión para agendar una cita', 'error')
        return redirect(url_for('login'))

    # Obtener servicio seleccionado
    servicio = Servicio.query.get_or_404(id_servicio)

    # Obtener empleados que realizan este servicio
    empleados_ids = db.session.query(EmpleadoServicio.id_empleado).filter_by(id_servicio=id_servicio).all()
    empleados_ids = [e[0] for e in empleados_ids]

    empleados = Empleado.query.filter(
        Empleado.id_empleado.in_(empleados_ids),
        Empleado.activo
    ).all()

    return render_template('citas/paso2_empleado.html', servicio=servicio, empleados=empleados)


@app.route('/citas/agendar/paso3/<int:id_servicio>/<int:id_empleado>')
def agendar_paso3(id_servicio, id_empleado):
    """Paso 3: Seleccionar fecha y hora"""
    if 'usuario_id' not in session:
        flash('Debes iniciar sesión para agendar una cita', 'error')
        return redirect(url_for('login'))

    servicio = Servicio.query.get_or_404(id_servicio)
    empleado = Empleado.query.get_or_404(id_empleado) if id_empleado > 0 else None

    # Fechas para el template
    hoy = datetime.now().strftime('%Y-%m-%d')
    max_fecha = (datetime.now() + timedelta(days=90)).strftime('%Y-%m-%d')

    return render_template('citas/paso3_fecha_hora.html',
                           servicio=servicio,
                           empleado=empleado,
                           hoy=hoy,
                           max_fecha=max_fecha)


@app.route('/citas/horarios-disponibles')
def horarios_disponibles():
    """API: Obtener horarios disponibles para una fecha y empleado"""
    fecha_str = request.args.get('fecha')
    id_empleado = request.args.get('id_empleado', type=int)
    id_servicio = request.args.get('id_servicio', type=int)

    if not all([fecha_str, id_servicio]):
        return jsonify({'error': 'Faltan parámetros'}), 400

    try:
        fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date()
    except BaseException:
        return jsonify({'error': 'Fecha inválida'}), 400

    # Si id_empleado es 0, seleccionar empleado aleatorio que haga el servicio
    if id_empleado == 0:
        empleados_ids = db.session.query(EmpleadoServicio.id_empleado).filter_by(id_servicio=id_servicio).all()
        empleados_ids = [e[0] for e in empleados_ids]
        if not empleados_ids:
            return jsonify({'horarios': []})
        id_empleado = random.choice(empleados_ids)

    # Obtener servicio para duración
    servicio = Servicio.query.get(id_servicio)
    if not servicio:
        return jsonify({'error': 'Servicio no encontrado'}), 404

    # Obtener día de la semana (0=Domingo, 1=Lunes, etc.)
    dia_semana = (fecha.weekday() + 1) % 7  # Convertir de Python (0=Lunes) a nuestra DB (0=Domingo)

    # Obtener horario del empleado para ese día
    horario = HorarioEmpleado.query.filter_by(
        id_empleado=id_empleado,
        dia_semana=dia_semana
    ).first()

    if not horario:
        return jsonify({'horarios': []})

    # Generar slots de tiempo disponibles
    horarios_disponibles = []
    hora_actual = datetime.combine(fecha, horario.hora_inicio)
    hora_fin = datetime.combine(fecha, horario.hora_fin)
    duracion = timedelta(minutes=servicio.duracion_minutos)

    while hora_actual + duracion <= hora_fin:
        # Verificar si ya hay una cita en este horario
        cita_existente = Cita.query.filter(
            Cita.id_empleado == id_empleado,
            Cita.fecha_hora_inicio < hora_actual + duracion,
            Cita.fecha_hora_fin > hora_actual,
            Cita.estado.in_(['pendiente_pago', 'confirmada', 'en_atencion'])
        ).first()

        if not cita_existente:
            horarios_disponibles.append({
                'hora': hora_actual.strftime('%H:%M'),
                'hora_fin': (hora_actual + duracion).strftime('%H:%M'),
                'disponible': True
            })

        hora_actual += timedelta(minutes=30)  # Intervalos de 30 minutos

    return jsonify({
        'horarios': horarios_disponibles,
        'id_empleado': id_empleado
    })


@app.route('/citas/agendar/paso4', methods=['POST'])
def agendar_paso4():
    """Paso 4: Confirmación y pago"""
    if 'usuario_id' not in session:
        flash('Debes iniciar sesión para agendar una cita', 'error')
        return redirect(url_for('login'))

    # Obtener datos del formulario
    id_servicio = request.form.get('id_servicio', type=int)
    id_empleado = request.form.get('id_empleado', type=int)
    fecha_str = request.form.get('fecha')
    hora_str = request.form.get('hora')

    if not all([id_servicio, fecha_str, hora_str]):
        flash('Datos incompletos', 'error')
        return redirect(url_for('agendar_paso1'))

    # Obtener información
    servicio = Servicio.query.get_or_404(id_servicio)
    empleado = Empleado.query.get_or_404(id_empleado) if id_empleado > 0 else None

    # Parsear fecha y hora
    try:
        fecha_hora_inicio = datetime.strptime(f"{fecha_str} {hora_str}", '%Y-%m-%d %H:%M')
        fecha_hora_fin = fecha_hora_inicio + timedelta(minutes=servicio.duracion_minutos)
    except BaseException:
        flash('Fecha u hora inválida', 'error')
        return redirect(url_for('agendar_paso1'))

    # Validar que la fecha sea futura
    if fecha_hora_inicio < datetime.now():
        flash('No puedes agendar citas en el pasado', 'error')
        return redirect(url_for('agendar_paso3', id_servicio=id_servicio, id_empleado=id_empleado or 0))

    return render_template('citas/paso4_confirmacion.html',
                           servicio=servicio,
                           empleado=empleado,
                           fecha_hora_inicio=fecha_hora_inicio,
                           fecha_hora_fin=fecha_hora_fin,
                           id_empleado=id_empleado or 0)


@app.route('/citas/confirmar', methods=['POST'])
def confirmar_cita():
    """
    Flujo completo de reserva usando ReservaService (6 pasos):
    1. Cliente inicia reserva
    2. Captura y valida datos del cliente (ClienteDTO)
    3. Genera link de pasarela de pago
    4. Procesa confirmación de pago
    5. Valida transacción (ServicePasarelaPago)
    6. Crea la cita en BD con estado confirmado y genera comprobante
    """
    if 'usuario_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401

    # ── Obtener datos del formulario ─────────────────────────────────────────
    id_servicio           = request.form.get('id_servicio', type=int)
    id_empleado           = request.form.get('id_empleado', type=int)
    fecha_hora_inicio_str = request.form.get('fecha_hora_inicio')
    fecha_hora_fin_str    = request.form.get('fecha_hora_fin')

    try:
        fecha_hora_inicio = datetime.strptime(fecha_hora_inicio_str, '%Y-%m-%d %H:%M:%S')
        fecha_hora_fin    = datetime.strptime(fecha_hora_fin_str,    '%Y-%m-%d %H:%M:%S')
    except (ValueError, TypeError):
        flash('Error en las fechas', 'error')
        return redirect(url_for('agendar_paso1'))

    # ── Obtener servicio ─────────────────────────────────────────────────────
    servicio = Servicio.query.get(id_servicio)
    if not servicio:
        flash('Servicio no encontrado', 'error')
        return redirect(url_for('agendar_paso1'))

    # ── Si empleado es 0, asignar uno aleatorio ──────────────────────────────
    if id_empleado == 0:
        empleados_ids = db.session.query(EmpleadoServicio.id_empleado)\
            .filter_by(id_servicio=id_servicio).all()
        empleados_ids = [e[0] for e in empleados_ids]
        if empleados_ids:
            id_empleado = random.choice(empleados_ids)
        else:
            flash('No hay empleados disponibles para este servicio', 'error')
            return redirect(url_for('agendar_paso1'))

    # ── Obtener datos del cliente desde la BD ────────────────────────────────
    cliente_db = Usuario.query.get(session['usuario_id'])
    if not cliente_db:
        flash('Cliente no encontrado', 'error')
        return redirect(url_for('agendar_paso1'))

    try:
        # ── PASO 1: Iniciar reserva ──────────────────────────────────────────
        reserva = ReservaService(
            servicio   = servicio.nombre_servicio,
            fecha_cita = fecha_hora_inicio
        )

        # ── PASO 2 y 3: Registrar datos del cliente y generar link pago ─────
        try:
            link_pago = reserva.registrar_datos_cliente(
                nombre   = cliente_db.nombre,
                telefono = cliente_db.telefono,
                correo   = cliente_db.email
            )
            logging.info(f"[RESERVA] Link de pasarela generado: {link_pago}")
        except ValueError as e:
            flash(f'Error en los datos del cliente: {str(e)}', 'error')
            return redirect(url_for('agendar_paso1'))

        # ── PASO 4 y 5: Confirmar pago (abono fijo $5.000 COP) ───────────────
        # En producción este paso llega por webhook de la pasarela.
        # Aquí lo simulamos con el abono registrado en el sistema.
        id_transaccion = f"TRX-{reserva.id_reserva}-{datetime.now().strftime('%Y%m%d%H%M%S')}"
        pago_exitoso   = reserva.recibir_confirmacion_pago(
            monto_pagado   = ReservaService.ABONO_REQUERIDO,
            transaction_id = id_transaccion
        )

        if not pago_exitoso:
            flash('Error al procesar el pago. Intenta de nuevo.', 'error')
            return redirect(url_for('agendar_paso1'))

        # ── PASO 6: Obtener comprobante y crear cita en BD ───────────────────
        comprobante = reserva.obtener_comprobante_cita()
        logging.info(
            f"[RESERVA] Comprobante generado: {comprobante['comprobante_id']} | "
            f"Cliente: {comprobante['cliente']['nombre']} | "
            f"Estado: {comprobante['estado']}"
        )

        # Generar token de gestión seguro (para link de reprogramación)
        token_gestion = secrets.token_urlsafe(16)

        # Crear la cita en PostgreSQL con el código del ReservaService
        nueva_cita = Cita(
            id_cliente        = session['usuario_id'],
            id_empleado       = id_empleado,
            id_servicio       = id_servicio,
            fecha_hora_inicio = fecha_hora_inicio,
            fecha_hora_fin    = fecha_hora_fin,
            monto_total       = Decimal(str(servicio.precio_total)),
            monto_abono       = Decimal(str(ReservaService.ABONO_REQUERIDO)),
            saldo_pendiente   = Decimal(str(servicio.precio_total)) - Decimal(str(ReservaService.ABONO_REQUERIDO)),
            estado            = 'pendiente_pago',
            reembolsado       = False,
            codigo_reserva    = reserva.id_reserva,
            token_gestion     = token_gestion,
            notas             = f"TRX: {id_transaccion} | Comprobante: {comprobante['comprobante_id']}",
            fecha_creacion    = datetime.now()
        )

        db.session.add(nueva_cita)
        db.session.commit()

        logging.info(
            f"[BD] Cita #{nueva_cita.id_cita} creada en PostgreSQL — "
            f"Reserva: {reserva.id_reserva} — Token gestión generado ✓"
        )

        # Notificar al cliente con el link de gestión
        try:
            link_gestion = url_for('gestionar_cita', token=token_gestion, _external=True)
            add_notificacion(
                session['usuario_id'],
                '📅 Cita agendada exitosamente',
                f'Tu cita de {servicio.nombre_servicio} para el '
                f'{fecha_hora_inicio.strftime("%d/%m/%Y a las %H:%M")} '
                f'fue registrada. Código: {reserva.id_reserva}. '
                f'Gestiona tu cita en: {link_gestion}',
                target=url_for('mis_citas')
            )
        except Exception:
            pass

        flash('¡Cita agendada exitosamente!', 'success')
        return redirect(url_for('cita_confirmada', codigo=reserva.id_reserva))

    except ReservaError as e:
        logging.error(f"[RESERVA] Error en el flujo: {e}")
        flash(f'Error en el proceso de reserva: {str(e)}', 'error')
        return redirect(url_for('agendar_paso1'))
    except Exception as e:
        db.session.rollback()
        logging.error(f"[RESERVA] Error inesperado: {e}")
        flash(f'Error al crear la cita: {str(e)}', 'error')
        return redirect(url_for('agendar_paso1'))


@app.route('/citas/confirmada/<codigo>')
def cita_confirmada(codigo):
    """Mostrar detalles de cita confirmada"""
    if 'usuario_id' not in session:
        flash('Debes iniciar sesión', 'error')
        return redirect(url_for('login'))

    cita = Cita.query.filter_by(codigo_reserva=codigo, id_cliente=session['usuario_id']).first_or_404()
    servicio = Servicio.query.get(cita.id_servicio)
    empleado = Empleado.query.get(cita.id_empleado)

    return render_template('citas/confirmada.html', cita=cita, servicio=servicio, empleado=empleado)


@app.route('/citas/mis-citas')
def mis_citas():
    """Ver mis citas agendadas"""
    if 'usuario_id' not in session:
        flash('Debes iniciar sesión', 'error')
        return redirect(url_for('login'))

    # Obtener citas futuras
    citas_futuras = db.session.query(Cita, Servicio, Empleado).join(
        Servicio, Cita.id_servicio == Servicio.id_servicio
    ).join(
        Empleado, Cita.id_empleado == Empleado.id_empleado
    ).filter(
        Cita.id_cliente == session['usuario_id'],
        Cita.fecha_hora_inicio >= datetime.now(),
        Cita.estado.in_(['pendiente_pago', 'confirmada'])
    ).order_by(Cita.fecha_hora_inicio).all()

    # Obtener citas pasadas
    citas_pasadas = db.session.query(Cita, Servicio, Empleado).join(
        Servicio, Cita.id_servicio == Servicio.id_servicio
    ).join(
        Empleado, Cita.id_empleado == Empleado.id_empleado
    ).filter(
        Cita.id_cliente == session['usuario_id'],
        Cita.fecha_hora_inicio < datetime.now()
    ).order_by(Cita.fecha_hora_inicio.desc()).limit(10).all()

    # Pasar función now para calcular tiempo restante en el template
    return render_template('citas/mis_citas.html',
                           citas_futuras=citas_futuras,
                           citas_pasadas=citas_pasadas,
                           now=datetime.now)


@app.route('/citas/cancelar/<int:id_cita>', methods=['POST'])
def cancelar_cita(id_cita):
    """
    Flujo completo de cancelación (basado en arquitectura PasarelaPagoService):
    1. Valida sesión y pertenencia de la cita.
    2. Verifica que la cita no esté ya cancelada.
    3. Evalúa tiempo restante >= 2 horas.
    4. Cancela la cita en la agenda.
    5. Procesa reembolso automático de $5.000 COP.
    6. Notifica al cliente por sistema (+ log tipo WhatsApp).
    7. Notifica a administradores.
    """
    if 'usuario_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401

    cita = Cita.query.filter_by(
        id_cita=id_cita,
        id_cliente=session['usuario_id']
    ).first()

    if not cita:
        return jsonify({'error': 'Cita no encontrada'}), 404

    # ── Paso 2: Verificar que no esté ya cancelada ──────────────────────────
    if cita.estado == 'cancelada':
        logging.warning(f"[CANCELACIÓN] Cita {cita.id_cita} ya fue cancelada anteriormente.")
        return jsonify({'error': 'Esta cita ya fue cancelada anteriormente'}), 400

    # ── Paso 3: Evaluar tiempo restante (Hora Cita - Hora Actual >= 2 horas) ─
    hora_actual      = datetime.now()
    tiempo_restante  = cita.fecha_hora_inicio - hora_actual
    horas_restantes  = tiempo_restante.total_seconds() / 3600

    if horas_restantes < Cita.HORAS_MINIMAS_CANCELACION if hasattr(Cita, 'HORAS_MINIMAS_CANCELACION') else 2:
        logging.warning(
            f"[CANCELACIÓN] Cita {cita.id_cita} — tiempo insuficiente: "
            f"{horas_restantes:.2f} hrs restantes."
        )
        return jsonify({
            'error': f'Debes cancelar con mínimo 2 horas de anticipación. '
                     f'Tiempo restante: {horas_restantes:.1f} hrs.'
        }), 400

    # ── Paso 4: Cancelar en la agenda ────────────────────────────────────────
    cita.estado      = 'cancelada'
    cita.reembolsado = False
    db.session.commit()
    logging.info(f"[CANCELACIÓN] Cita {cita.id_cita} cancelada exitosamente en la agenda.")

    # ── Paso 5: Procesar reembolso automático del abono ($5.000 COP) ─────────
    id_transaccion    = f"TRX-CITA-{cita.id_cita}-{cita.codigo_reserva}"
    reembolso_exitoso = PasarelaPagoService.procesar_reembolso(
        id_transaccion=id_transaccion,
        monto=PasarelaPagoService.REEMBOLSO_MONTO,
        moneda=PasarelaPagoService.MONEDA
    )

    if reembolso_exitoso:
        cita.reembolsado = True
        db.session.commit()
        mensaje_cliente = (
            f"Hola {cita.cliente.nombre if cita.cliente else 'clienta'}, "
            f"tu cita del {cita.fecha_hora_inicio.strftime('%d/%m/%Y a las %H:%M')} "
            f"ha sido cancelada con éxito. "
            f"Se ha procesado el reembolso de $5.000 COP a tu método de pago."
        )
        logging.info(
            f"[PASARELA DE PAGO] Reembolso exitoso — TRX: {id_transaccion} — "
            f"Monto: ${PasarelaPagoService.REEMBOLSO_MONTO:,.0f} COP"
        )
    else:
        mensaje_cliente = (
            f"Hola {cita.cliente.nombre if cita.cliente else 'clienta'}, "
            f"tu cita del {cita.fecha_hora_inicio.strftime('%d/%m/%Y a las %H:%M')} "
            f"fue cancelada. Se generó una orden de devolución manual de $5.000 COP "
            f"con administración. Te contactaremos pronto."
        )
        logging.error(
            f"[PASARELA DE PAGO] Reembolso fallido — TRX: {id_transaccion}. "
            f"Se requiere gestión manual."
        )

    # ── Paso 6: Notificar al cliente (sistema + log WhatsApp) ─────────────────
    try:
        # Notificación interna del sistema
        add_notificacion(
            cita.id_cliente,
            '🔔 Cita cancelada' + (' — Reembolso procesado ✓' if reembolso_exitoso else ' — Reembolso pendiente'),
            mensaje_cliente,
            target=url_for('mis_citas')
        )
        # Log simulando envío por WhatsApp (integrar con API de WhatsApp en producción)
        logging.info(
            f"[NOTIFICACIÓN - EMAIL]    Enviando a "
            f"{cita.cliente.nombre if cita.cliente else cita.id_cliente}: '{mensaje_cliente}'"
        )
        logging.info(
            f"[NOTIFICACIÓN - WHATSAPP] Enviando a "
            f"{cita.cliente.telefono if cita.cliente else 'N/A'}: '{mensaje_cliente}'"
        )
    except Exception as e:
        logging.error(f"[NOTIFICACIÓN] Error al notificar cliente: {e}")

    # ── Paso 7: Notificar a administradores ───────────────────────────────────
    try:
        estado_reembolso = "✓ Reembolso procesado" if reembolso_exitoso else "⚠ Reembolso pendiente — requiere gestión manual"
        admins = Usuario.query.filter_by(tipo_usuario='admin').all()
        for a in admins:
            add_notificacion(
                a.id,
                f'Cita #{cita.id_cita} cancelada — {estado_reembolso}',
                (
                    f"Cliente: {cita.cliente.nombre if cita.cliente else cita.id_cliente} | "
                    f"Cita: {cita.fecha_hora_inicio.strftime('%d/%m/%Y %H:%M')} | "
                    f"Servicio: {cita.servicio.nombre_servicio if cita.servicio else 'N/A'} | "
                    f"TRX: {id_transaccion} | "
                    f"Reembolso $5.000 COP: {'procesado' if reembolso_exitoso else 'PENDIENTE MANUAL'}"
                ),
                target=url_for('admin_citas') + f'?estado=cancelada&cliente_id={cita.id_cliente}'
            )
    except Exception as e:
        logging.error(f"[NOTIFICACIÓN] Error al notificar admins: {e}")

    return jsonify({
        'success':   True,
        'message':   'Cita cancelada exitosamente',
        'reembolso': reembolso_exitoso,
        'detalle':   mensaje_cliente
    })

@app.route('/citas/pagar/<int:id_cita>', methods=['GET', 'POST'])
def cliente_pagos_registrar(id_cita):
    """Registrar pago para una cita por el cliente"""
    if 'usuario_id' not in session:
        flash('Debes iniciar sesión para realizar un pago', 'error')
        return redirect(url_for('login'))

    cita = Cita.query.get_or_404(id_cita)
    
    # Ensure the appointment belongs to the logged in user
    if cita.id_cliente != session['usuario_id']:
        flash('No tienes permiso para pagar esta cita', 'error')
        return redirect(url_for('mis_citas'))

    cliente = Usuario.query.get(cita.id_cliente)
    servicio = Servicio.query.get(cita.id_servicio)

    # Verificar que no tenga ya un pago o no esté cancelada
    if cita.pago:
        flash('Esta cita ya tiene un pago registrado', 'error')
        return redirect(url_for('mis_citas'))
    
    if cita.estado == 'cancelada':
        flash('No puedes pagar una cita cancelada', 'error')
        return redirect(url_for('mis_citas'))

    if request.method == 'POST':
        monto = request.form.get('monto', type=float)
        metodo = request.form.get('metodo_pago', 'efectivo')
        referencia = request.form.get('referencia', '').strip() or None
        notas = request.form.get('notas', '').strip() or None

        if not monto or monto <= 0:
            flash('El monto debe ser mayor a 0', 'error')
            return redirect(url_for('cliente_pagos_registrar', id_cita=id_cita))

        metodos_validos = ['efectivo', 'tarjeta', 'transferencia', 'nequi', 'daviplata']
        if metodo not in metodos_validos:
            flash('Método de pago inválido', 'error')
            return redirect(url_for('cliente_pagos_registrar', id_cita=id_cita))

        nuevo_pago = Pago(
            id_cita=id_cita,
            monto=Decimal(str(monto)),
            metodo_pago=metodo,
            estado_pago='completado',
            referencia=referencia,
            notas=notas
        )
        db.session.add(nuevo_pago)

        # Actualizar estado de cita y saldo
        cita.monto_abono = (cita.monto_abono or Decimal('0')) + Decimal(str(monto))
        cita.saldo_pendiente = (cita.monto_total or Decimal('0')) - cita.monto_abono
        if cita.saldo_pendiente <= 0:
            cita.estado = 'completada'
            cita.saldo_pendiente = Decimal('0')

        db.session.commit()
        
        # Notificar a administradores
        try:
            admins = Usuario.query.filter_by(tipo_usuario='admin').all()
            for a in admins:
                add_notificacion(
                    a.id,
                    'Pago registrado por cliente',
                    f'Pago de ${monto:,.0f} registrado por el cliente {cliente.nombre} para la cita #{cita.id_cita}',
                    target=url_for('admin_pagos')
                )
        except Exception:
            pass

        flash(f'Pago de ${monto:,.0f} registrado exitosamente', 'success')
        return redirect(url_for('mis_citas'))

    # GET — formulario
    metodos = ['efectivo', 'tarjeta', 'transferencia', 'nequi', 'daviplata']
    return render_template('citas/cliente_pagos_form.html',
                           cita=cita, cliente=cliente,
                           servicio=servicio, metodos=metodos)

# ============================================================================
# RUTAS PANEL ADMIN - GESTIÓN DE EMPLEADOS
# ============================================================================


@app.route('/admin/empleados')
@admin_required
def admin_empleados():
    """Listar todos los empleados"""
    empleados = Empleado.query.order_by(Empleado.nombre).all()
    return render_template('admin/empleados.html', empleados=empleados)


@app.route('/admin/empleados/crear', methods=['GET', 'POST'])
@admin_required
def admin_empleados_crear():
    """Crear nuevo empleado"""
    if request.method == 'POST':
        nombre = request.form.get('nombre')
        especialidad = request.form.get('especialidad')
        servicios_ids = request.form.getlist('servicios')

        if not nombre:
            flash('El nombre es obligatorio', 'error')
            return redirect(url_for('admin_empleados_crear'))

        # Crear empleado
        nuevo_empleado = Empleado(
            nombre=nombre,
            especialidad=especialidad,
            activo=True
        )

        db.session.add(nuevo_empleado)
        db.session.flush()  # Para obtener el ID

        # Asignar servicios
        for id_servicio in servicios_ids:
            empleado_servicio = EmpleadoServicio(
                id_empleado=nuevo_empleado.id_empleado,
                id_servicio=int(id_servicio)
            )
            db.session.add(empleado_servicio)

        db.session.commit()
        flash(f'Empleado {nombre} creado exitosamente', 'success')
        return redirect(url_for('admin_empleados'))

    # GET: Mostrar formulario
    servicios = Servicio.query.filter_by(activo=True).all()
    return render_template('admin/empleados_form.html', empleado=None, servicios=servicios)


@app.route('/admin/empleados/editar/<int:id_empleado>', methods=['GET', 'POST'])
@admin_required
def admin_empleados_editar(id_empleado):
    """Editar empleado existente"""
    empleado = Empleado.query.get_or_404(id_empleado)

    if request.method == 'POST':
        empleado.nombre = request.form.get('nombre')
        empleado.especialidad = request.form.get('especialidad')
        empleado.activo = request.form.get('activo') == 'on'

        # Actualizar servicios
        servicios_ids = request.form.getlist('servicios')

        # Eliminar relaciones existentes
        EmpleadoServicio.query.filter_by(id_empleado=id_empleado).delete()

        # Agregar nuevas relaciones
        for id_servicio in servicios_ids:
            empleado_servicio = EmpleadoServicio(
                id_empleado=id_empleado,
                id_servicio=int(id_servicio)
            )
            db.session.add(empleado_servicio)

        db.session.commit()
        flash(f'Empleado {empleado.nombre} actualizado exitosamente', 'success')
        return redirect(url_for('admin_empleados'))

    # GET: Mostrar formulario
    servicios = Servicio.query.filter_by(activo=True).all()
    servicios_empleado = [es.id_servicio for es in EmpleadoServicio.query.filter_by(id_empleado=id_empleado).all()]

    return render_template('admin/empleados_form.html',
                           empleado=empleado,
                           servicios=servicios,
                           servicios_empleado=servicios_empleado)


@app.route('/admin/empleados/clientes-afectados/<int:id_empleado>', methods=['GET'])
@admin_required
def admin_empleados_clientes_afectados(id_empleado):
    """API: Retorna clientes con citas futuras asignadas a este empleado."""
    empleado = Empleado.query.get_or_404(id_empleado)
    citas = db.session.query(Cita, Usuario, Servicio)\
        .join(Usuario, Cita.id_cliente == Usuario.id)\
        .join(Servicio, Cita.id_servicio == Servicio.id_servicio)\
        .filter(
            Cita.id_empleado == id_empleado,
            Cita.fecha_hora_inicio >= datetime.now(),
            Cita.estado.in_(['pendiente_pago', 'confirmada'])
        ).order_by(Cita.fecha_hora_inicio).all()

    afectados = []
    for cita, cliente, servicio in citas:
        afectados.append({
            'id_cita':    cita.id_cita,
            'codigo':     cita.codigo_reserva,
            'cliente':    cliente.nombre,
            'telefono':   cliente.telefono,
            'email':      cliente.email,
            'servicio':   servicio.nombre_servicio,
            'fecha':      cita.fecha_hora_inicio.strftime('%d/%m/%Y'),
            'hora':       cita.fecha_hora_inicio.strftime('%H:%M'),
        })

    return jsonify({
        'empleado':  empleado.nombre,
        'total':     len(afectados),
        'afectados': afectados
    })


@app.route('/admin/empleados/eliminar/<int:id_empleado>', methods=['POST'])
@admin_required
def admin_empleados_eliminar(id_empleado):
    """
    Eliminar empleado.
    Si tiene citas futuras, las desasigna (id_empleado = NULL)
    y las deja en 'pendiente_pago' para reprogramación manual.
    El admin debe haber confirmado explícitamente este paso.
    """
    empleado = Empleado.query.get_or_404(id_empleado)
    confirmado = request.form.get('confirmado') == 'true'

    # Buscar citas futuras activas
    citas_futuras = Cita.query.filter(
        Cita.id_empleado == id_empleado,
        Cita.fecha_hora_inicio >= datetime.now(),
        Cita.estado.in_(['pendiente_pago', 'confirmada'])
    ).all()

    if citas_futuras and not confirmado:
        return jsonify({
            'success':  False,
            'requiere_confirmacion': True,
            'message':  f'El empleado tiene {len(citas_futuras)} cita(s) pendiente(s). '
                        f'Confirma para desasignarlas y programar reprogramación manual.'
        }), 409

    nombre = empleado.nombre

    # Desasignar citas futuras → quedan sin empleado, estado pendiente_pago
    citas_desasignadas = 0
    for cita in citas_futuras:
        cita.id_empleado = None
        cita.estado      = 'pendiente_pago'
        # Notificar al cliente
        try:
            add_notificacion(
                cita.id_cliente,
                '⚠️ Tu cita necesita reprogramación',
                f'La especialista asignada a tu cita del '
                f'{cita.fecha_hora_inicio.strftime("%d/%m/%Y a las %H:%M")} '
                f'ya no está disponible. El equipo Rossmix te contactará para reprogramarla.',
                target=url_for('mis_citas')
            )
        except Exception:
            pass
        citas_desasignadas += 1
        logging.info(
            f"[ELIMINAR EMPLEADO] Cita #{cita.id_cita} desasignada "
            f"— cliente notificado para reprogramación manual."
        )

    db.session.delete(empleado)
    db.session.commit()

    # Notificar a admins sobre citas que requieren reprogramación
    if citas_desasignadas > 0:
        try:
            admins = Usuario.query.filter_by(tipo_usuario='admin').all()
            for a in admins:
                add_notificacion(
                    a.id,
                    f'⚠️ {citas_desasignadas} cita(s) requieren reprogramación manual',
                    f'Se eliminó a {nombre}. '
                    f'{citas_desasignadas} cita(s) quedaron sin especialista asignada. '
                    f'Revisa la sección de Citas para reprogramarlas.',
                    target=url_for('admin_citas') + '?estado=pendiente_pago'
                )
        except Exception:
            pass

    return jsonify({
        'success':            True,
        'message':            f'Empleado {nombre} eliminado exitosamente.',
        'citas_desasignadas': citas_desasignadas
    })

# ============================================================================
# RUTAS PANEL ADMIN - GESTIÓN DE SERVICIOS
# ============================================================================


@app.route('/admin/servicios')
@admin_required
def admin_servicios():
    """Listar todos los servicios"""
    servicios = Servicio.query.order_by(Servicio.nombre_servicio).all()
    return render_template('admin/servicios.html', servicios=servicios)


@app.route('/admin/servicios/crear', methods=['GET', 'POST'])
@admin_required
def admin_servicios_crear():
    """Crear nuevo servicio"""
    if request.method == 'POST':
        nombre = request.form.get('nombre_servicio')
        descripcion = request.form.get('descripcion')
        precio = request.form.get('precio_total')
        duracion = request.form.get('duracion_minutos')

        if not all([nombre, precio, duracion]):
            flash('Todos los campos son obligatorios', 'error')
            return redirect(url_for('admin_servicios_crear'))

        # Crear servicio
        nuevo_servicio = Servicio(
            nombre_servicio=nombre,
            descripcion=descripcion,
            precio_total=Decimal(precio),
            duracion_minutos=int(duracion),
            activo=True
        )

        db.session.add(nuevo_servicio)
        db.session.commit()

        flash(f'Servicio {nombre} creado exitosamente', 'success')
        return redirect(url_for('admin_servicios'))

    return render_template('admin/servicios_form.html', servicio=None)


@app.route('/admin/servicios/editar/<int:id_servicio>', methods=['GET', 'POST'])
@admin_required
def admin_servicios_editar(id_servicio):
    """Editar servicio existente"""
    servicio = Servicio.query.get_or_404(id_servicio)

    if request.method == 'POST':
        servicio.nombre_servicio = request.form.get('nombre_servicio')
        servicio.descripcion = request.form.get('descripcion')
        servicio.precio_total = Decimal(request.form.get('precio_total'))
        servicio.duracion_minutos = int(request.form.get('duracion_minutos'))
        servicio.activo = request.form.get('activo') == 'on'

        db.session.commit()
        flash(f'Servicio {servicio.nombre_servicio} actualizado exitosamente', 'success')
        return redirect(url_for('admin_servicios'))

    return render_template('admin/servicios_form.html', servicio=servicio)


@app.route('/admin/servicios/eliminar/<int:id_servicio>', methods=['POST'])
@admin_required
def admin_servicios_eliminar(id_servicio):
    """Eliminar servicio"""
    servicio = Servicio.query.get_or_404(id_servicio)

    # Verificar si tiene citas futuras
    citas_futuras = Cita.query.filter(
        Cita.id_servicio == id_servicio,
        Cita.fecha_hora_inicio >= datetime.now(),
        Cita.estado.in_(['pendiente_pago', 'confirmada'])
    ).count()

    if citas_futuras > 0:
        return jsonify({
            'success': False,
            'message': f'No se puede eliminar. El servicio tiene {citas_futuras} cita(s) pendiente(s)'
        }), 400

    nombre = servicio.nombre_servicio
    db.session.delete(servicio)
    db.session.commit()

    return jsonify({
        'success': True,
        'message': f'Servicio {nombre} eliminado exitosamente'
    })

# ============================================================================
# RUTAS PANEL ADMIN - GESTIÓN DE CLIENTES
# ============================================================================


@app.route('/admin/clientes')
@admin_required
def admin_clientes():
    """Listar todos los clientes"""
    clientes = Usuario.query.filter_by(tipo_usuario='cliente').order_by(Usuario.nombre).all()
    # Añadir conteo de citas canceladas por cliente
    for c in clientes:
        try:
            c.citas_canceladas = Cita.query.filter_by(id_cliente=c.id, estado='cancelada').count()
        except Exception:
            c.citas_canceladas = 0
    return render_template('admin/clientes.html', clientes=clientes, filter_label='Todos')


@app.route('/admin/clientes/hoy')
@admin_required
def admin_clientes_hoy():
    """Listar clientes registrados hoy"""
    from sqlalchemy import func

    hoy = datetime.now().date()
    clientes = Usuario.query.filter(
        Usuario.tipo_usuario == 'cliente',
        func.date(Usuario.fecha_registro) == hoy
    ).order_by(Usuario.nombre).all()

    return render_template(
        'admin/clientes.html',
        clientes=clientes,
        view_title='Clientes registrados hoy',
        filter_label='Hoy'
    )


@app.route('/admin/clientes/editar/<int:id_cliente>', methods=['GET', 'POST'])
@admin_required
def admin_clientes_editar(id_cliente):
    """Editar cliente"""
    cliente = Usuario.query.get_or_404(id_cliente)

    if request.method == 'POST':
        cliente.nombre = request.form.get('nombre')
        cliente.email = request.form.get('email')
        cliente.telefono = request.form.get('telefono')
        cliente.activo = request.form.get('activo') == 'on'

        # Cambiar contraseña solo si se proporciona una nueva
        nueva_password = request.form.get('nueva_password')
        if nueva_password:
            cliente.password = generate_password_hash(nueva_password)

        db.session.commit()
        flash(f'Cliente {cliente.nombre} actualizado exitosamente', 'success')
        return redirect(url_for('admin_clientes'))

    return render_template('admin/clientes_form.html', cliente=cliente)


@app.route('/admin/clientes/eliminar/<int:id_cliente>', methods=['POST'])
@admin_required
def admin_clientes_eliminar(id_cliente):
    """Eliminar cliente"""
    cliente = Usuario.query.get_or_404(id_cliente)

    # Verificar si tiene citas futuras
    citas_futuras = Cita.query.filter(
        Cita.id_cliente == id_cliente,
        Cita.fecha_hora_inicio >= datetime.now(),
        Cita.estado.in_(['pendiente_pago', 'confirmada'])
    ).count()

    if citas_futuras > 0:
        return jsonify({
            'success': False,
            'message': f'No se puede eliminar. El cliente tiene {citas_futuras} cita(s) pendiente(s)'
        }), 400

    nombre = cliente.nombre
    db.session.delete(cliente)
    db.session.commit()

    return jsonify({
        'success': True,
        'message': f'Cliente {nombre} eliminado exitosamente'
    })

# ============================================================================
# RUTAS PANEL ADMIN - GESTIÓN DE HORARIOS
# ============================================================================


@app.route('/admin/horarios')
@admin_required
def admin_horarios():
    """Listar horarios de todos los empleados"""
    empleados = Empleado.query.filter_by(activo=True).all()
    dias_semana = {
        0: 'Domingo', 1: 'Lunes', 2: 'Martes', 3: 'Miércoles',
        4: 'Jueves', 5: 'Viernes', 6: 'Sábado'
    }
    return render_template('admin/horarios.html', empleados=empleados, dias_semana=dias_semana)


@app.route('/admin/horarios/crear/<int:id_empleado>', methods=['GET', 'POST'])
@admin_required
def admin_horarios_crear(id_empleado):
    """Crear horario para empleado"""
    empleado = Empleado.query.get_or_404(id_empleado)

    if request.method == 'POST':
        dia_semana = int(request.form.get('dia_semana'))
        hora_inicio_str = request.form.get('hora_inicio')
        hora_fin_str = request.form.get('hora_fin')

        # Validar que no exista ya un horario para ese día
        horario_existente = HorarioEmpleado.query.filter_by(
            id_empleado=id_empleado,
            dia_semana=dia_semana
        ).first()

        if horario_existente:
            flash('Ya existe un horario para este empleado en ese día', 'error')
            return redirect(url_for('admin_horarios_crear', id_empleado=id_empleado))

        # Convertir strings a time
        hora_inicio = datetime.strptime(hora_inicio_str, '%H:%M').time()
        hora_fin = datetime.strptime(hora_fin_str, '%H:%M').time()

        if hora_inicio >= hora_fin:
            flash('La hora de inicio debe ser menor que la hora de fin', 'error')
            return redirect(url_for('admin_horarios_crear', id_empleado=id_empleado))

        # Crear horario
        nuevo_horario = HorarioEmpleado(
            id_empleado=id_empleado,
            dia_semana=dia_semana,
            hora_inicio=hora_inicio,
            hora_fin=hora_fin
        )

        db.session.add(nuevo_horario)
        db.session.commit()

        flash(f'Horario creado exitosamente para {empleado.nombre}', 'success')
        return redirect(url_for('admin_horarios'))

    # GET
    dias_semana = {
        0: 'Domingo', 1: 'Lunes', 2: 'Martes', 3: 'Miércoles',
        4: 'Jueves', 5: 'Viernes', 6: 'Sábado'
    }
    return render_template('admin/horarios_form.html', empleado=empleado, horario=None, dias_semana=dias_semana)


@app.route('/admin/horarios/editar/<int:id_horario>', methods=['GET', 'POST'])
@admin_required
def admin_horarios_editar(id_horario):
    """Editar horario"""
    horario = HorarioEmpleado.query.get_or_404(id_horario)

    if request.method == 'POST':
        hora_inicio_str = request.form.get('hora_inicio')
        hora_fin_str = request.form.get('hora_fin')

        # Convertir strings a time
        hora_inicio = datetime.strptime(hora_inicio_str, '%H:%M').time()
        hora_fin = datetime.strptime(hora_fin_str, '%H:%M').time()

        if hora_inicio >= hora_fin:
            flash('La hora de inicio debe ser menor que la hora de fin', 'error')
            return redirect(url_for('admin_horarios_editar', id_horario=id_horario))

        horario.hora_inicio = hora_inicio
        horario.hora_fin = hora_fin

        db.session.commit()
        flash('Horario actualizado exitosamente', 'success')
        return redirect(url_for('admin_horarios'))

    # GET
    dias_semana = {
        0: 'Domingo', 1: 'Lunes', 2: 'Martes', 3: 'Miércoles',
        4: 'Jueves', 5: 'Viernes', 6: 'Sábado'
    }
    return render_template(
        'admin/horarios_form.html',
        empleado=horario.empleado,
        horario=horario,
        dias_semana=dias_semana)


@app.route('/admin/horarios/eliminar/<int:id_horario>', methods=['POST'])
@admin_required
def admin_horarios_eliminar(id_horario):
    """Eliminar horario"""
    horario = HorarioEmpleado.query.get_or_404(id_horario)

    db.session.delete(horario)
    db.session.commit()

    return jsonify({
        'success': True,
        'message': 'Horario eliminado exitosamente'
    })

# ============================================================================
# RUTAS PANEL ADMIN - GESTIÓN DE CITAS
# ============================================================================


@app.route('/admin/citas')
@admin_required
def admin_citas():
    """Listar todas las citas"""
    # Filtros
    estado = request.args.get('estado', 'todas')
    fecha_desde = request.args.get('fecha_desde')
    fecha_hasta = request.args.get('fecha_hasta')

    query = db.session.query(Cita, Usuario, Empleado, Servicio).join(
        Usuario, Cita.id_cliente == Usuario.id
    ).outerjoin(
        Empleado, Cita.id_empleado == Empleado.id_empleado
    ).join(
        Servicio, Cita.id_servicio == Servicio.id_servicio
    )

    # Aplicar filtros
    if estado != 'todas':
        query = query.filter(Cita.estado == estado)

    # Filtrar por cliente si se pasa cliente_id
    cliente_id = request.args.get('cliente_id')
    if cliente_id:
        try:
            cid = int(cliente_id)
            query = query.filter(Cita.id_cliente == cid)
        except ValueError:
            pass

    if fecha_desde:
        fecha_desde_dt = datetime.strptime(fecha_desde, '%Y-%m-%d')
        query = query.filter(Cita.fecha_hora_inicio >= fecha_desde_dt)

    if fecha_hasta:
        fecha_hasta_dt = datetime.strptime(fecha_hasta, '%Y-%m-%d') + timedelta(days=1)
        query = query.filter(Cita.fecha_hora_inicio < fecha_hasta_dt)

    citas = query.order_by(Cita.fecha_hora_inicio.desc()).all()

    # Si se filtró por cliente, obtener objeto para mostrar en la cabecera
    cliente_filtrado = None
    cliente_id = request.args.get('cliente_id')
    if cliente_id:
        try:
            cliente_filtrado = Usuario.query.get(int(cliente_id))
        except Exception:
            cliente_filtrado = None

    return render_template('admin/citas.html', citas=citas, estado_filtro=estado, cliente_filtrado=cliente_filtrado)


@app.route('/admin/citas/cambiar-estado/<int:id_cita>', methods=['POST'])
@admin_required
def admin_citas_cambiar_estado(id_cita):
    """Cambiar estado de una cita"""
    cita = Cita.query.get_or_404(id_cita)
    nuevo_estado = request.form.get('estado')

    estados_validos = ['pendiente_pago', 'confirmada', 'en_atencion', 'completada', 'cancelada', 'no_asistio']

    if nuevo_estado not in estados_validos:
        return jsonify({'success': False, 'message': 'Estado inválido'}), 400

    cita.estado = nuevo_estado
    db.session.commit()

    return jsonify({
        'success': True,
        'message': f'Estado de cita actualizado a {nuevo_estado}'
    })


@app.route('/admin/citas/reasignar-empleado/<int:id_cita>', methods=['GET', 'POST'])
@admin_required
def admin_citas_reasignar_empleado(id_cita):
    """
    GET  — devuelve la lista de empleados disponibles para el servicio de la cita.
    POST — asigna el nuevo empleado y confirma la cita.
    """
    cita     = Cita.query.get_or_404(id_cita)
    servicio = Servicio.query.get(cita.id_servicio)

    if request.method == 'GET':
        # Empleados que realizan este servicio
        empleados_ids = db.session.query(EmpleadoServicio.id_empleado)\
            .filter_by(id_servicio=cita.id_servicio).all()
        empleados_ids = [e[0] for e in empleados_ids]
        empleados = Empleado.query.filter(
            Empleado.id_empleado.in_(empleados_ids),
            Empleado.activo == True
        ).order_by(Empleado.nombre).all()
        return jsonify({
            'id_cita':  id_cita,
            'servicio': servicio.nombre_servicio if servicio else 'N/A',
            'cliente':  cita.cliente.nombre if cita.cliente else 'N/A',
            'fecha':    cita.fecha_hora_inicio.strftime('%d/%m/%Y a las %H:%M'),
            'empleados': [
                {'id': e.id_empleado, 'nombre': e.nombre, 'especialidad': e.especialidad or ''}
                for e in empleados
            ]
        })

    # POST — reasignar
    nuevo_empleado_id = request.form.get('id_empleado', type=int)
    if not nuevo_empleado_id:
        return jsonify({'success': False, 'message': 'Selecciona un empleado'}), 400

    empleado = Empleado.query.get(nuevo_empleado_id)
    if not empleado:
        return jsonify({'success': False, 'message': 'Empleado no encontrado'}), 404

    cita.id_empleado = nuevo_empleado_id
    # Si estaba cancelada o sin asignar, reactivar como pendiente_pago
    if cita.estado in ['cancelada', 'no_asistio']:
        cita.estado = 'pendiente_pago'
    else:
        cita.estado = 'confirmada'
    db.session.commit()

    logging.info(
        f"[REASIGNAR] Cita #{id_cita} reasignada a {empleado.nombre} — "
        f"estado → confirmada."
    )

    # Notificar al cliente
    try:
        add_notificacion(
            cita.id_cliente,
            '✅ Tu cita fue reasignada y confirmada',
            f'Tu cita de {servicio.nombre_servicio if servicio else "servicio"} '
            f'del {cita.fecha_hora_inicio.strftime("%d/%m/%Y a las %H:%M")} '
            f'fue asignada a {empleado.nombre}. ¡Te esperamos!',
            target=url_for('mis_citas')
        )
    except Exception:
        pass

    return jsonify({
        'success':         True,
        'message':         f'Cita #{id_cita} reasignada a {empleado.nombre} y confirmada.',
        'empleado_nombre': empleado.nombre,
        'nuevo_estado':    'confirmada'
    })

# ============================================================================
# RUTAS PANEL ADMIN - GESTIÓN DE PAGOS
# ============================================================================


@app.route('/admin/pagos')
@admin_required
def admin_pagos():
    """Listar todos los pagos registrados"""
    pagos = db.session.query(Pago, Cita, Usuario, Servicio)\
        .join(Cita, Pago.id_cita == Cita.id_cita)\
        .join(Usuario, Cita.id_cliente == Usuario.id)\
        .join(Servicio, Cita.id_servicio == Servicio.id_servicio)\
        .order_by(Pago.fecha_pago.desc()).all()

    return render_template('admin/pagos.html', pagos=pagos)


@app.route('/admin/pagos/registrar/<int:id_cita>', methods=['GET', 'POST'])
@admin_required
def admin_pagos_registrar(id_cita):
    """Registrar pago para una cita"""
    cita = Cita.query.get_or_404(id_cita)
    cliente = Usuario.query.get(cita.id_cliente)
    servicio = Servicio.query.get(cita.id_servicio)

    # Verificar que no tenga ya un pago
    if cita.pago:
        flash('Esta cita ya tiene un pago registrado', 'error')
        return redirect(url_for('admin_pagos'))

    if request.method == 'POST':
        monto = request.form.get('monto', type=float)
        metodo = request.form.get('metodo_pago', 'efectivo')
        referencia = request.form.get('referencia', '').strip() or None
        notas = request.form.get('notas', '').strip() or None

        if not monto or monto <= 0:
            flash('El monto debe ser mayor a 0', 'error')
            return redirect(url_for('admin_pagos_registrar', id_cita=id_cita))

        metodos_validos = ['efectivo', 'tarjeta', 'transferencia', 'nequi', 'daviplata']
        if metodo not in metodos_validos:
            flash('Método de pago inválido', 'error')
            return redirect(url_for('admin_pagos_registrar', id_cita=id_cita))

        nuevo_pago = Pago(
            id_cita=id_cita,
            monto=Decimal(str(monto)),
            metodo_pago=metodo,
            estado_pago='completado',
            referencia=referencia,
            notas=notas
        )
        db.session.add(nuevo_pago)

        # Actualizar estado de cita y saldo
        cita.monto_abono = Decimal(str(monto))
        cita.saldo_pendiente = (cita.monto_total or Decimal('0')) - Decimal(str(monto))
        if cita.saldo_pendiente <= 0:
            cita.estado = 'completada'

        db.session.commit()
        # Notificar al cliente
        try:
            add_notificacion(
                cita.id_cliente,
                'Pago registrado',
                f'Se registró un pago de ${monto:,.0f} para tu cita. Saldo pendiente: ${cita.saldo_pendiente:,.0f}',
                target=url_for('mis_citas')
            )
        except Exception:
            pass

        # Notificar a administradores
        try:
            admins = Usuario.query.filter_by(tipo_usuario='admin').all()
            for a in admins:
                add_notificacion(
                    a.id,
                    'Pago registrado',
                    f'Pago de ${monto:,.0f} registrado para la cita #{cita.id_cita} del cliente {cliente.nombre}',
                    target=url_for('admin_pagos')
                )
        except Exception:
            pass

        flash(f'Pago de ${monto:,.0f} registrado exitosamente', 'success')
        return redirect(url_for('admin_pagos'))

    # GET — formulario
    metodos = ['efectivo', 'tarjeta', 'transferencia', 'nequi', 'daviplata']
    return render_template('admin/pagos_form.html',
                           cita=cita, cliente=cliente,
                           servicio=servicio, metodos=metodos)


@app.route('/admin/pagos/eliminar/<int:id_pago>', methods=['POST'])
@admin_required
def admin_pagos_eliminar(id_pago):
    """Eliminar un pago (reembolso)"""
    pago = Pago.query.get_or_404(id_pago)
    pago.cita.reembolsado = True
    pago.cita.estado = 'cancelada'
    db.session.delete(pago)
    db.session.commit()
    return jsonify({'success': True, 'message': 'Pago eliminado y cita marcada como reembolsada'})
@app.route('/admin/exportar/<tipo>/<periodo>')
@admin_required
def admin_exportar_excel(tipo, periodo):
    if tipo not in ['citas', 'pagos', 'empleados', 'servicios', 'clientes', 'horarios']:
        flash('Tipo de exportación no válido.', 'error')
        return redirect(url_for('dashboard_admin'))

    hoy = datetime.now()
    if periodo == 'diario':
        fecha_inicio = hoy.replace(hour=0, minute=0, second=0, microsecond=0)
    elif periodo == 'semana':
        fecha_inicio = hoy - timedelta(days=7)
    elif periodo == 'mes':
        fecha_inicio = hoy.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    elif periodo == 'ano':
        fecha_inicio = hoy.replace(month=1, day=1, hour=0, minute=0, second=0, microsecond=0)
    else:
        fecha_inicio = datetime(1900, 1, 1)

    wb = openpyxl.Workbook()
    ws = wb.active

    if tipo == 'citas':
        ws.title = "Citas"
        ws.append(["ID", "Código", "Cliente", "Servicio", "Empleado", "Fecha/Hora", "Estado",
                   "Total ($)", "Abono ($)", "Saldo Pendiente ($)", "¿Abonado?", "Fecha Creación"])
        for cell in ws[1]:
            cell.fill = __import__('openpyxl').styles.PatternFill(start_color="F2B5D4", end_color="F2B5D4", fill_type="solid")
            cell.font  = __import__('openpyxl').styles.Font(bold=True)

        query = Cita.query.filter(Cita.fecha_creacion >= fecha_inicio).all()
        for c in query:
            cli = Usuario.query.get(c.id_cliente)
            srv = Servicio.query.get(c.id_servicio)
            emp = Empleado.query.get(c.id_empleado) if c.id_empleado else None
            abono     = float(c.monto_abono or 0)
            total     = float(c.monto_total or 0)
            saldo     = float(c.saldo_pendiente or 0)
            abonado   = 'Sí ✓' if abono > 0 else 'No ✗'
            ws.append([
                c.id_cita,
                c.codigo_reserva,
                cli.nombre if cli else 'N/A',
                srv.nombre_servicio if srv else 'N/A',
                emp.nombre if emp else 'Sin asignar',
                c.fecha_hora_inicio.strftime('%d/%m/%Y %H:%M'),
                c.estado.replace('_', ' ').upper(),
                total,
                abono,
                saldo,
                abonado,
                c.fecha_creacion.strftime('%d/%m/%Y %H:%M')
            ])
            # Colorear fila según si abonó
            fila = ws.max_row
            fill_si  = __import__('openpyxl').styles.PatternFill(start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
            fill_no  = __import__('openpyxl').styles.PatternFill(start_color="F8D7DA", end_color="F8D7DA", fill_type="solid")
            ws.cell(row=fila, column=11).fill = fill_si if abono > 0 else fill_no
            ws.cell(row=fila, column=11).font = __import__('openpyxl').styles.Font(
                bold=True, color="155724" if abono > 0 else "721C24"
            )
            
    elif tipo == 'pagos':
        ws.title = "Pagos"
        ws.append(["ID Pago", "Código Cita", "Monto", "Método", "Estado", "Fecha"])
        query = Pago.query.filter(Pago.fecha_pago >= fecha_inicio).all()
        for p in query:
            cita = Cita.query.get(p.id_cita)
            ws.append([p.id_pago, cita.codigo_reserva if cita else '', float(p.monto), p.metodo_pago, p.estado_pago, p.fecha_pago.strftime('%Y-%m-%d %H:%M')])
            
    elif tipo == 'empleados':
        ws.title = "Empleados"
        ws.append(["ID", "Nombre", "Email", "Especialidad", "Estado"])
        query = Usuario.query.filter_by(tipo_usuario='empleado').filter(Usuario.fecha_registro >= fecha_inicio).all()
        for e in query:
            ws.append([e.id, e.nombre, e.email, getattr(e, 'especialidad', ''), "Activo" if getattr(e, 'activo', True) else "Inactivo"])
            
    elif tipo == 'clientes':
        ws.title = "Clientes"
        ws.append(["ID", "Nombre", "Email", "Teléfono", "Fecha Registro"])
        query = Usuario.query.filter_by(tipo_usuario='cliente').filter(Usuario.fecha_registro >= fecha_inicio).all()
        for c in query:
            ws.append([c.id, c.nombre, c.email, c.telefono, c.fecha_registro.strftime('%Y-%m-%d %H:%M')])
            
    elif tipo == 'servicios':
        ws.title = "Servicios"
        ws.append(["ID", "Nombre", "Descripción", "Precio", "Duración"])
        query = Servicio.query.all() # No date filter for services
        for s in query:
            ws.append([s.id_servicio, s.nombre_servicio, s.descripcion, float(s.precio_total), s.duracion_minutos])
            
    elif tipo == 'horarios':
        ws.title = "Horarios"
        ws.append(["ID Horario", "Empleado", "Día", "Hora Inicio", "Hora Fin"])
        query = HorarioEmpleado.query.all()
        for h in query:
            emp = Empleado.query.get(h.id_empleado)
            ws.append([h.id_horario, emp.nombre if emp else '', h.dia_semana, h.hora_inicio.strftime('%H:%M'), h.hora_fin.strftime('%H:%M')])

    # Ajuste automático de ancho de columnas
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value and len(str(cell.value)) > max_len:
                    max_len = len(str(cell.value))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        as_attachment=True,
        download_name=f"export_{tipo}_{periodo}.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

# ============================================================================
# RUTAS PANEL ADMIN - PAGOS POR CONFIRMAR
# ============================================================================

@app.route('/admin/pagos-por-confirmar')
@admin_required
def admin_pagos_confirmar():
    """Listar citas en estado pendiente_pago para que el admin confirme el abono"""
    citas = db.session.query(Cita, Usuario, Servicio, Empleado)\
        .join(Usuario, Cita.id_cliente == Usuario.id)\
        .join(Servicio, Cita.id_servicio == Servicio.id_servicio)\
        .outerjoin(Empleado, Cita.id_empleado == Empleado.id_empleado)\
        .filter(Cita.estado == 'pendiente_pago')\
        .order_by(Cita.fecha_hora_inicio.asc()).all()

    return render_template('admin/pagos_confirmar.html', citas=citas)


@app.route('/admin/pagos-por-confirmar/aceptar/<int:id_cita>', methods=['POST'])
@admin_required
def admin_aceptar_pago(id_cita):
    """Confirmar el pago de una cita: cambia estado a 'confirmada' y notifica a la clienta"""
    cita = Cita.query.get_or_404(id_cita)

    if cita.estado != 'pendiente_pago':
        return jsonify({'success': False, 'message': 'Esta cita ya fue procesada'}), 400

    cita.estado = 'confirmada'
    db.session.commit()

    # Notificar a la clienta
    try:
        add_notificacion(
            cita.id_cliente,
            '¡Cita Confirmada! 🎉',
            f'Tu pago fue verificado. Tu cita del {cita.fecha_hora_inicio.strftime("%d/%m/%Y a las %H:%M")} está confirmada. ¡Te esperamos!',
            target=url_for('mis_citas')
        )
    except Exception:
        pass

    return jsonify({'success': True, 'message': 'Pago aceptado y cita confirmada'})


# ============================================================================
# RUTAS — GESTIÓN Y REPROGRAMACIÓN DE CITAS (SistemaGestionCitas)
# ============================================================================

@app.route('/citas/gestionar/<token>')
def gestionar_cita(token: str):
    """
    Paso 1: Cliente accede desde el link de gestión recibido por WhatsApp/email.
    Muestra las opciones: Reprogramar o Cancelar.
    """
    cita = Cita.query.filter_by(token_gestion=token).first()
    if not cita:
        flash('Enlace de gestión inválido o expirado.', 'error')
        return redirect(url_for('index'))

    try:
        datos = SistemaGestionCitas.acceder_a_gestion(token, cita)
    except ReprogramacionError as e:
        flash(str(e), 'error')
        return redirect(url_for('index'))

    servicio = Servicio.query.get(cita.id_servicio)
    empleado = Empleado.query.get(cita.id_empleado) if cita.id_empleado else None

    return render_template(
        'citas/gestionar_cita.html',
        cita     = cita,
        datos    = datos,
        servicio = servicio,
        empleado = empleado,
        token    = token
    )


@app.route('/citas/reprogramar/<token>')
def reprogramar_cita_form(token: str):
    """
    Pasos 2, 3 y 4: Valida la política de reprogramación
    y muestra el calendario de disponibilidad.
    """
    cita = Cita.query.filter_by(token_gestion=token).first()
    if not cita:
        flash('Enlace inválido o expirado.', 'error')
        return redirect(url_for('index'))

    # Paso 2 y 3: Validar política de tiempo
    try:
        SistemaGestionCitas.validar_politica_reprogramacion(cita)
    except ReprogramacionError as e:
        flash(str(e), 'error')
        return redirect(url_for('gestionar_cita', token=token))

    # Paso 4: Obtener disponibilidad
    disponibilidad = SistemaGestionCitas.obtener_agenda_disponible(
        id_servicio         = cita.id_servicio,
        id_empleado_actual  = cita.id_empleado
    )

    servicio = Servicio.query.get(cita.id_servicio)

    return render_template(
        'citas/reprogramar.html',
        cita           = cita,
        servicio       = servicio,
        disponibilidad = disponibilidad,
        token          = token
    )


@app.route('/citas/reprogramar/<token>/confirmar', methods=['POST'])
def reprogramar_cita_confirmar(token: str):
    """
    Pasos 5 y 6: Ejecuta la reprogramación, mantiene el abono
    y notifica al cliente.
    """
    cita = Cita.query.filter_by(token_gestion=token).first()
    if not cita:
        return jsonify({'success': False, 'message': 'Enlace inválido'}), 404

    nueva_fecha_str  = request.form.get('nueva_fecha')
    nuevo_empleado   = request.form.get('id_empleado', type=int)

    if not nueva_fecha_str:
        return jsonify({'success': False, 'message': 'Fecha no proporcionada'}), 400

    try:
        nueva_fecha = datetime.strptime(nueva_fecha_str, '%Y-%m-%d %H:%M')
    except ValueError:
        return jsonify({'success': False, 'message': 'Formato de fecha inválido'}), 400

    # Validar política antes de ejecutar
    try:
        SistemaGestionCitas.validar_politica_reprogramacion(cita)
    except ReprogramacionError as e:
        return jsonify({'success': False, 'message': str(e)}), 400

    # Pasos 5 y 6: Ejecutar reprogramación
    try:
        resultado = SistemaGestionCitas.ejecutar_reprogramacion(
            cita_orm          = cita,
            nueva_fecha_hora  = nueva_fecha,
            nuevo_id_empleado = nuevo_empleado
        )
    except Exception as e:
        db.session.rollback()
        logging.error(f"[REPROGRAMACIÓN] Error: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500

    # Notificar al admin
    try:
        admins = Usuario.query.filter_by(tipo_usuario='admin').all()
        for a in admins:
            add_notificacion(
                a.id,
                f'🔄 Cita #{cita.id_cita} reprogramada',
                f'Cliente: {cita.cliente.nombre if cita.cliente else "N/A"} | '
                f'Nueva fecha: {resultado["nueva_fecha"]} | '
                f'Profesional: {resultado["profesional"]}',
                target=url_for('admin_citas')
            )
    except Exception:
        pass

    # Notificación interna al cliente
    try:
        add_notificacion(
            cita.id_cliente,
            '🔄 Cita reprogramada con éxito',
            f'Tu cita fue reprogramada para el {resultado["nueva_fecha"]} '
            f'con {resultado["profesional"]}. '
            f'Abono de ${resultado["abono_mantenido"]:,.0f} COP conservado.',
            target=url_for('mis_citas')
        )
    except Exception:
        pass

    return jsonify({
        'success':  True,
        'message':  '¡Cita reprogramada exitosamente!',
        'resultado': resultado
    })


@app.route('/citas/disponibilidad/<int:id_servicio>')
def api_disponibilidad(id_servicio: int):
    """API: Retorna slots disponibles para un servicio dado."""
    id_empleado = request.args.get('id_empleado', type=int)
    disponibilidad = SistemaGestionCitas.obtener_agenda_disponible(
        id_servicio        = id_servicio,
        id_empleado_actual = id_empleado
    )
    return jsonify(disponibilidad)


# ============================================================================
# RUTAS PANEL ADMIN — AGENDA DIARIA (SistemaAgendaDiaria)
# ============================================================================

@app.route('/admin/agenda-diaria')
@app.route('/admin/agenda-diaria/<fecha_str>')
@admin_required
def admin_agenda_diaria(fecha_str: str = None):
    """
    Pasos 1 y 2: Vista de Agenda Diaria.
    Carga citas del día desde PostgreSQL, construye SistemaAgendaDiaria
    y devuelve la cuadrícula agrupada por profesional al template.
    """
    # Parsear fecha — por defecto hoy
    try:
        fecha = datetime.strptime(fecha_str, '%Y-%m-%d').date() if fecha_str else date.today()
    except ValueError:
        fecha = date.today()

    # Cargar citas del día desde la BD (incluyendo confirmadas y en_atencion)
    from sqlalchemy import func
    citas_orm = db.session.query(Cita)\
        .filter(
            func.date(Cita.fecha_hora_inicio) == fecha,
            Cita.estado.in_(['pendiente_pago', 'confirmada', 'en_atencion', 'completada'])
        )\
        .order_by(Cita.fecha_hora_inicio).all()

    # Construir sistema y cuadrícula usando SistemaAgendaDiaria
    sistema = SistemaAgendaDiaria()
    sistema.cargar_desde_bd(citas_orm)
    cuadricula = sistema.obtener_cuadrilla_agenda_diaria(fecha)

    logging.info(
        f"[AGENDA DIARIA] Fecha: {fecha} | "
        f"Citas: {len(citas_orm)} | "
        f"Profesionales: {len(cuadricula)}"
    )

    # Estadísticas del día
    stats_dia = {
        'total':       len(citas_orm),
        'programadas': sum(1 for c in citas_orm if c.estado in ['pendiente_pago', 'confirmada']),
        'en_atencion': sum(1 for c in citas_orm if c.estado == 'en_atencion'),
        'completadas': sum(1 for c in citas_orm if c.estado == 'completada'),
        'ingresos':    sum(float(c.monto_total or 0) for c in citas_orm if c.estado == 'completada'),
    }

    # Días navegables (semana actual)
    fecha_prev = (datetime.combine(fecha, time.min) - timedelta(days=1)).date()
    fecha_next = (datetime.combine(fecha, time.min) + timedelta(days=1)).date()

    # Citas del mes completo para el calendario
    from sqlalchemy import func as sqlfunc2
    primer_dia_mes = fecha.replace(day=1)
    if fecha.month == 12:
        ultimo_dia_mes = fecha.replace(year=fecha.year+1, month=1, day=1) - timedelta(days=1)
    else:
        ultimo_dia_mes = fecha.replace(month=fecha.month+1, day=1) - timedelta(days=1)

    citas_mes = db.session.query(
        sqlfunc2.date(Cita.fecha_hora_inicio).label('dia'),
        sqlfunc2.count(Cita.id_cita).label('total')
    ).filter(
        Cita.fecha_hora_inicio >= datetime.combine(primer_dia_mes, time.min),
        Cita.fecha_hora_inicio <= datetime.combine(ultimo_dia_mes, time.max),
        Cita.estado.in_(['pendiente_pago', 'confirmada', 'en_atencion', 'completada'])
    ).group_by(sqlfunc2.date(Cita.fecha_hora_inicio)).all()

    # Diccionario {dia_str: total}
    dias_con_citas = {str(r.dia): r.total for r in citas_mes}

    metodos_pago = [m.value for m in MetodoPagoSaldo]

    return render_template(
        'admin/agenda_diaria.html',
        cuadricula    = cuadricula,
        fecha         = fecha,
        fecha_prev    = fecha_prev,
        fecha_next    = fecha_next,
        stats_dia     = stats_dia,
        metodos_pago  = metodos_pago,
        dias_con_citas = dias_con_citas,
        primer_dia_mes = primer_dia_mes,
        ultimo_dia_mes = ultimo_dia_mes,
    )


@app.route('/admin/agenda-diaria/en-atencion/<int:id_cita>', methods=['POST'])
@admin_required
def admin_marcar_en_atencion(id_cita: int):
    """
    Paso 3: El cliente llegó — cambia estado a 'en_atencion' en PostgreSQL.
    Usa SistemaAgendaDiaria.marcar_en_atencion() para la lógica de negocio.
    """
    cita = Cita.query.get_or_404(id_cita)

    # Validar con SistemaAgendaDiaria
    sistema = SistemaAgendaDiaria()
    sistema.cargar_desde_bd([cita])
    resultado = sistema.marcar_en_atencion(str(id_cita))

    if not resultado:
        return jsonify({
            'success': False,
            'message': f'No se puede marcar en atención. Estado actual: {cita.estado}'
        }), 400

    # Persistir en PostgreSQL
    cita.estado = 'en_atencion'
    db.session.commit()

    logging.info(
        f"[AGENDA] Cita #{id_cita} → 'en_atencion' | "
        f"Cliente: {cita.cliente.nombre if cita.cliente else 'N/A'}"
    )

    # Notificar al cliente
    try:
        add_notificacion(
            cita.id_cliente,
            '💅 ¡Tu turno ha comenzado!',
            f'Tu servicio de {cita.servicio.nombre_servicio if cita.servicio else "belleza"} '
            f'está en curso. ¡Disfrútalo!',
            target=url_for('mis_citas')
        )
    except Exception:
        pass

    return jsonify({
        'success': True,
        'message': f'Cita #{id_cita} marcada como En Atención',
        'estado':  'en_atencion'
    })


@app.route('/admin/agenda-diaria/liquidar/<int:id_cita>', methods=['POST'])
@admin_required
def admin_liquidar_cita(id_cita: int):
    """
    Pasos 4 y 5: Finaliza el servicio, cobra el saldo pendiente
    y marca la cita como 'completada' usando SistemaAgendaDiaria.
    """
    cita         = Cita.query.get_or_404(id_cita)
    metodo_str   = request.form.get('metodo_pago', 'Efectivo')

    # Mapear string → MetodoPagoSaldo
    metodo_map = {m.value: m for m in MetodoPagoSaldo}
    metodo     = metodo_map.get(metodo_str, MetodoPagoSaldo.EFECTIVO)

    # Validar y ejecutar con SistemaAgendaDiaria
    sistema = SistemaAgendaDiaria()
    sistema.cargar_desde_bd([cita])

    try:
        ticket = sistema.liquidar_y_completar_cita(str(id_cita), metodo)
    except InvalidOperationError as e:
        return jsonify({'success': False, 'message': str(e)}), 400
    except ValueError as e:
        return jsonify({'success': False, 'message': str(e)}), 404

    # Persistir en PostgreSQL
    saldo_cobrado       = Decimal(str(ticket['saldo_cobrado']))
    cita.estado         = 'completada'
    cita.monto_abono    = (cita.monto_abono or Decimal('0')) + saldo_cobrado
    cita.saldo_pendiente= Decimal('0')
    db.session.commit()

    # Registrar pago en tabla Pago
    try:
        metodo_bd_map = {
            'Efectivo':      'efectivo',
            'Transferencia': 'transferencia',
            'Nequi':         'nequi',
            'Daviplata':     'daviplata',
            'Tarjeta':       'tarjeta',
        }
        nuevo_pago = Pago(
            id_cita     = id_cita,
            monto       = saldo_cobrado,
            metodo_pago = metodo_bd_map.get(metodo_str, 'efectivo'),
            estado_pago = 'completado',
            notas       = f'Saldo liquidado en recepción — {metodo_str}'
        )
        db.session.add(nuevo_pago)
        db.session.commit()
    except Exception as e:
        logging.error(f"[AGENDA] Error al registrar pago: {e}")

    # Notificar al cliente
    try:
        add_notificacion(
            cita.id_cliente,
            '✅ Servicio completado',
            f'Tu servicio de {cita.servicio.nombre_servicio if cita.servicio else "belleza"} '
            f'fue completado. Saldo cobrado: ${ticket["saldo_cobrado"]:,.0f} COP '
            f'({ticket["metodo_pago"]}). ¡Gracias por visitarnos!',
            target=url_for('mis_citas')
        )
    except Exception:
        pass

    logging.info(
        f"[AGENDA] Cita #{id_cita} COMPLETADA | "
        f"Total: ${ticket['precio_total']:,.0f} | "
        f"Abono: ${ticket['abono_previo']:,.0f} | "
        f"Saldo cobrado: ${ticket['saldo_cobrado']:,.0f} ({ticket['metodo_pago']})"
    )

    return jsonify({
        'success':       True,
        'message':       f'Cita #{id_cita} completada exitosamente',
        'estado':        'completada',
        'precio_total':  ticket['precio_total'],
        'abono_previo':  ticket['abono_previo'],
        'saldo_cobrado': ticket['saldo_cobrado'],
        'metodo_pago':   ticket['metodo_pago']
    })


# ============================================================================
# RUTA: MARCAR CITA COMO "NO_ASISTIO" — Abono no reembolsable
# ============================================================================

@app.route('/admin/agenda-diaria/no-asistio/<int:id_cita>', methods=['POST'])
@admin_required
def admin_marcar_no_asistio(id_cita: int):
    """
    El admin marca la cita como 'no_asistio' una vez pasada la hora.
    - Estado cambia a 'no_asistio'
    - El abono de $5.000 COP NO es reembolsable
    - Se notifica al cliente con la opción de reagendar
    """
    cita = Cita.query.get_or_404(id_cita)

    estados_validos = ['confirmada', 'pendiente_pago', 'en_atencion']
    if cita.estado not in estados_validos:
        return jsonify({
            'success': False,
            'message': f'La cita ya está en estado: {cita.estado}'
        }), 400

    # Validar que la hora ya pasó
    if cita.fecha_hora_inicio > datetime.now():
        return jsonify({
            'success': False,
            'message': 'La cita aún no ha comenzado. No se puede marcar como no asistió.'
        }), 400

    cita.estado      = 'no_asistio'
    cita.reembolsado = False   # Abono no reembolsable
    db.session.commit()

    logging.info(
        f"[NO ASISTIO] Cita #{id_cita} — Cliente: "
        f"{cita.cliente.nombre if cita.cliente else 'N/A'} — "
        f"Abono ${float(cita.monto_abono or 5000):,.0f} COP NO reembolsable."
    )

    # Notificar al cliente con opción de reagendar
    try:
        link_reagendar = url_for('reagendar_no_asistio', id_cita=id_cita, _external=False)
        add_notificacion(
            cita.id_cliente,
            '⚠️ No asististe a tu cita',
            f'No registramos tu asistencia a la cita del '
            f'{cita.fecha_hora_inicio.strftime("%d/%m/%Y a las %H:%M")} '
            f'({cita.servicio.nombre_servicio if cita.servicio else "servicio"}). '
            f'El abono de $5.000 COP no es reembolsable según nuestra política. '
            f'Puedes reagendar tu cita y el abono se aplicará como crédito.',
            target=link_reagendar
        )
    except Exception as e:
        logging.error(f"[NO ASISTIO] Error al notificar: {e}")

    # Notificar admins
    try:
        admins = Usuario.query.filter_by(tipo_usuario='admin').all()
        for a in admins:
            add_notificacion(
                a.id,
                f'📋 Cita #{id_cita} marcada como No Asistió',
                f'Cliente: {cita.cliente.nombre if cita.cliente else "N/A"} — '
                f'Servicio: {cita.servicio.nombre_servicio if cita.servicio else "N/A"} — '
                f'Abono $5.000 COP no reembolsable.',
                target=url_for('admin_citas') + '?estado=no_asistio'
            )
    except Exception:
        pass

    return jsonify({
        'success': True,
        'message': f'Cita #{id_cita} marcada como No Asistió. Abono no reembolsable.',
        'estado':  'no_asistio',
        'abono':   float(cita.monto_abono or 5000)
    })


@app.route('/citas/reagendar-no-asistio/<int:id_cita>', methods=['GET', 'POST'])
def reagendar_no_asistio(id_cita: int):
    """
    El cliente puede reagendar una cita 'no_asistio'.
    El abono previo ($5.000 COP) se aplica como crédito a la nueva cita.
    """
    if 'usuario_id' not in session:
        flash('Debes iniciar sesión', 'error')
        return redirect(url_for('login'))

    cita_original = Cita.query.filter_by(
        id_cita=id_cita, id_cliente=session['usuario_id']
    ).first_or_404()

    if cita_original.estado != 'no_asistio':
        flash('Esta cita no está disponible para reagendar.', 'error')
        return redirect(url_for('mis_citas'))

    servicio = Servicio.query.get(cita_original.id_servicio)

    if request.method == 'GET':
        # Mostrar disponibilidad para reagendar
        disponibilidad = SistemaGestionCitas.obtener_agenda_disponible(
            id_servicio        = cita_original.id_servicio,
            id_empleado_actual = cita_original.id_empleado
        )
        return render_template(
            'citas/reagendar_no_asistio.html',
            cita_original = cita_original,
            servicio      = servicio,
            disponibilidad= disponibilidad,
            abono_credito = float(cita_original.monto_abono or 5000)
        )

    # POST — confirmar reagendamiento
    nueva_fecha_str = request.form.get('nueva_fecha')
    nuevo_empleado  = request.form.get('id_empleado', type=int)

    if not nueva_fecha_str:
        return jsonify({'success': False, 'message': 'Selecciona una fecha'}), 400

    try:
        nueva_fecha = datetime.strptime(nueva_fecha_str, '%Y-%m-%d %H:%M')
    except ValueError:
        return jsonify({'success': False, 'message': 'Formato de fecha inválido'}), 400

    empleado_id = nuevo_empleado or cita_original.id_empleado
    empleado    = Empleado.query.get(empleado_id) if empleado_id else None
    fecha_fin   = nueva_fecha + timedelta(
        minutes=servicio.duracion_minutos if servicio else 60
    )
    abono_credito = float(cita_original.monto_abono or 5000)
    token_nuevo   = secrets.token_urlsafe(16)

    # Crear nueva cita con el abono aplicado como crédito
    nueva_cita = Cita(
        id_cliente        = session['usuario_id'],
        id_empleado       = empleado_id,
        id_servicio       = cita_original.id_servicio,
        fecha_hora_inicio = nueva_fecha,
        fecha_hora_fin    = fecha_fin,
        monto_total       = cita_original.monto_total,
        monto_abono       = Decimal(str(abono_credito)),   # Crédito del abono anterior
        saldo_pendiente   = (cita_original.monto_total or Decimal('0')) - Decimal(str(abono_credito)),
        estado            = 'pendiente_pago',
        reembolsado       = False,
        codigo_reserva    = f"RE-{secrets.token_urlsafe(4).upper()}",
        token_gestion     = token_nuevo,
        notas             = f"Reagendada desde cita #{id_cita} (no asistió). "
                            f"Abono ${abono_credito:,.0f} COP aplicado como crédito.",
        fecha_creacion    = datetime.now()
    )

    try:
        db.session.add(nueva_cita)
        db.session.commit()

        logging.info(
            f"[REAGENDAR] Nueva cita #{nueva_cita.id_cita} creada desde cita #{id_cita} "
            f"— Abono ${abono_credito:,.0f} COP aplicado como crédito."
        )

        # Notificar al cliente
        add_notificacion(
            session['usuario_id'],
            '📅 Cita reagendada con crédito aplicado',
            f'Tu nueva cita de {servicio.nombre_servicio if servicio else "servicio"} '
            f'para el {nueva_fecha.strftime("%d/%m/%Y a las %H:%M")} fue registrada. '
            f'Se aplicó un crédito de ${abono_credito:,.0f} COP de tu abono anterior.',
            target=url_for('mis_citas')
        )

        return jsonify({
            'success':       True,
            'message':       '¡Cita reagendada exitosamente con tu crédito aplicado!',
            'nueva_cita_id': nueva_cita.id_cita,
            'codigo':        nueva_cita.codigo_reserva,
            'abono_credito': abono_credito
        })

    except Exception as e:
        db.session.rollback()
        logging.error(f"[REAGENDAR] Error: {e}")
        return jsonify({'success': False, 'message': str(e)}), 500


# ============================================================================
# API: ESTADO DE CITA (para polling en tiempo real)
# ============================================================================

@app.route('/citas/estado/<int:id_cita>')
def cita_estado(id_cita):
    """Devuelve el estado actual de una cita (usado por polling del frontend)"""
    if 'usuario_id' not in session:
        return jsonify({'error': 'No autorizado'}), 401
    cita = Cita.query.get_or_404(id_cita)
    if cita.id_cliente != session['usuario_id'] and session.get('tipo_usuario') != 'admin':
        return jsonify({'error': 'No autorizado'}), 403
    return jsonify({'estado': cita.estado, 'id_cita': cita.id_cita})


# ============================================================================
# DESCARGA PDF DETALLES DE CITA
# ============================================================================

@app.route('/citas/descargar/<int:id_cita>')
def descargar_cita_pdf(id_cita):
    """
    Genera y descarga el comprobante PDF usando obtener_comprobante_cita()
    del ReservaService para poblar los datos del documento.
    """
    if 'usuario_id' not in session:
        flash('Debes iniciar sesión', 'error')
        return redirect(url_for('login'))

    cita = Cita.query.get_or_404(id_cita)

    if cita.id_cliente != session['usuario_id'] and session.get('tipo_usuario') != 'admin':
        flash('No tienes permiso para esta acción', 'error')
        return redirect(url_for('mis_citas'))

    cliente  = Usuario.query.get(cita.id_cliente)
    servicio = Servicio.query.get(cita.id_servicio)
    empleado = Empleado.query.get(cita.id_empleado) if cita.id_empleado else None

    # ── Reconstruir ReservaService para obtener comprobante estructurado ─────
    # Usamos los datos de la cita ya guardada para simular el estado CONFIRMADA
    try:
        reserva_temp = ReservaService(
            servicio   = servicio.nombre_servicio if servicio else 'Servicio',
            fecha_cita = cita.fecha_hora_inicio
        )
        # Registrar datos del cliente (pasa a PENDIENTE_PAGO)
        reserva_temp.registrar_datos_cliente(
            nombre   = cliente.nombre   if cliente else 'N/A',
            telefono = cliente.telefono if cliente else 'N/A',
            correo   = cliente.email    if cliente else 'N/A'
        )
        # Extraer TRX de las notas si existe, si no generar uno de referencia
        id_trx = 'N/A'
        if cita.notas:
            for parte in cita.notas.split('|'):
                if 'TRX:' in parte:
                    id_trx = parte.replace('TRX:', '').strip()
                    break

        # Confirmar pago para llegar a estado CONFIRMADA
        reserva_temp.recibir_confirmacion_pago(
            monto_pagado   = ReservaService.ABONO_REQUERIDO,
            transaction_id = id_trx
        )
        # Sobreescribir id_reserva con el código real de la cita
        object.__setattr__(reserva_temp, 'id_reserva', cita.codigo_reserva) \
            if hasattr(reserva_temp, '__dataclass_fields__') else \
            setattr(reserva_temp, 'id_reserva', cita.codigo_reserva)

        comprobante = reserva_temp.obtener_comprobante_cita()
        logging.info(
            f"[PDF] Comprobante generado: {comprobante['comprobante_id']} | "
            f"Cliente: {comprobante['cliente']['nombre']}"
        )
    except Exception as e:
        logging.error(f"[PDF] Error al reconstruir ReservaService: {e}")
        # Fallback: construir comprobante manual con datos de la BD
        comprobante = {
            "comprobante_id": f"DOC-{cita.codigo_reserva}",
            "estado":         EstadoReserva.CONFIRMADA.value if cita.estado == 'confirmada' else cita.estado,
            "detalles_servicio": {
                "servicio":   servicio.nombre_servicio if servicio else 'N/A',
                "fecha_hora": cita.fecha_hora_inicio.strftime("%Y-%m-%d %H:%M"),
            },
            "cliente": {
                "nombre":   cliente.nombre   if cliente else 'N/A',
                "telefono": cliente.telefono if cliente else 'N/A',
                "correo":   cliente.email    if cliente else 'N/A',
            },
            "pago": {
                "monto_abono":    f"${float(cita.monto_abono or 0):,.0f} COP",
                "id_transaccion": id_trx if 'id_trx' in dir() else 'N/A',
            }
        }

    # ── Construir PDF con datos del comprobante ──────────────────────────────
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=A4,
        rightMargin=2*cm, leftMargin=2*cm,
        topMargin=2*cm,   bottomMargin=2*cm
    )

    PINK   = colors.HexColor('#c41e3a')
    LPINK  = colors.HexColor('#fff0f6')
    GREY   = colors.HexColor('#888888')
    BLACK  = colors.HexColor('#1a1a1a')
    GREEN  = colors.HexColor('#16a34a')

    estado_map = {
        'confirmada':     ('✓ CONFIRMADA',        GREEN),
        'pendiente_pago': ('⏳ PENDIENTE DE PAGO', colors.HexColor('#d97706')),
        'completada':     ('✓ COMPLETADA',         GREEN),
        'cancelada':      ('✗ CANCELADA',          colors.HexColor('#dc2626')),
    }
    estado_texto, estado_color = estado_map.get(
        cita.estado, (cita.estado.upper(), GREY)
    )

    title_s  = ParagraphStyle('t',  fontName='Helvetica-Bold', fontSize=22, textColor=PINK,  alignment=TA_CENTER, spaceAfter=4)
    sub_s    = ParagraphStyle('s',  fontName='Helvetica',      fontSize=11, textColor=GREY,  alignment=TA_CENTER, spaceAfter=2)
    sec_s    = ParagraphStyle('sc', fontName='Helvetica-Bold', fontSize=11, textColor=PINK,  spaceAfter=4)
    footer_s = ParagraphStyle('f',  fontName='Helvetica',      fontSize=9,  textColor=GREY,  alignment=TA_CENTER)

    def make_table(data, col_widths=[5*cm, 11*cm]):
        t = Table(data, colWidths=col_widths)
        t.setStyle(TableStyle([
            ('FONTNAME',        (0,0), (0,-1), 'Helvetica-Bold'),
            ('FONTNAME',        (1,0), (1,-1), 'Helvetica'),
            ('FONTSIZE',        (0,0), (-1,-1), 10),
            ('TEXTCOLOR',       (0,0), (0,-1),  GREY),
            ('TEXTCOLOR',       (1,0), (1,-1),  BLACK),
            ('ROWBACKGROUNDS',  (0,0), (-1,-1), [colors.white, LPINK]),
            ('TOPPADDING',      (0,0), (-1,-1), 8),
            ('BOTTOMPADDING',   (0,0), (-1,-1), 8),
            ('LEFTPADDING',     (0,0), (-1,-1), 10),
        ]))
        return t

    elements = []

    # Cabecera
    elements.append(Paragraph('Rossmix', title_s))
    elements.append(Paragraph('Salón de Belleza & Uñas', sub_s))
    elements.append(Paragraph('instagram.com/rossmiix | WhatsApp: +57 317 382 4030',
        ParagraphStyle('ig', fontName='Helvetica', fontSize=9, textColor=PINK, alignment=TA_CENTER)))
    elements.append(Spacer(1, 0.3*cm))
    elements.append(HRFlowable(width='100%', thickness=2, color=PINK))
    elements.append(Spacer(1, 0.4*cm))

    # Título y estado
    elements.append(Paragraph('COMPROBANTE DE CITA', ParagraphStyle(
        'dt', fontName='Helvetica-Bold', fontSize=16, textColor=BLACK, alignment=TA_CENTER, spaceAfter=2)))
    elements.append(Paragraph(estado_texto, ParagraphStyle(
        'es', fontName='Helvetica-Bold', fontSize=13, textColor=estado_color, alignment=TA_CENTER, spaceAfter=6)))
    elements.append(Spacer(1, 0.5*cm))

    # ── Sección: Información de Reserva (datos del ReservaService) ───────────
    elements.append(Paragraph('INFORMACIÓN DE RESERVA', sec_s))
    elements.append(HRFlowable(width='100%', thickness=1, color=colors.HexColor('#ffd6e8')))
    elements.append(Spacer(1, 0.2*cm))
    elements.append(make_table([
        ['ID Comprobante',  comprobante['comprobante_id']],
        ['Código Reserva',  cita.codigo_reserva],
        ['ID Transacción',  comprobante['pago']['id_transaccion']],
        ['Estado Reserva',  comprobante['estado']],
        ['Fecha Emisión',   datetime.now().strftime('%d/%m/%Y %H:%M')],
    ]))
    elements.append(Spacer(1, 0.5*cm))

    # ── Sección: Datos de la Clienta (del ClienteDTO) ────────────────────────
    elements.append(Paragraph('DATOS DE LA CLIENTA', sec_s))
    elements.append(HRFlowable(width='100%', thickness=1, color=colors.HexColor('#ffd6e8')))
    elements.append(Spacer(1, 0.2*cm))
    elements.append(make_table([
        ['Nombre',    comprobante['cliente']['nombre']],
        ['Correo',    comprobante['cliente']['correo']],
        ['Teléfono',  comprobante['cliente']['telefono']],
    ]))
    elements.append(Spacer(1, 0.5*cm))

    # ── Sección: Detalles del Servicio ───────────────────────────────────────
    elements.append(Paragraph('DETALLES DEL SERVICIO', sec_s))
    elements.append(HRFlowable(width='100%', thickness=1, color=colors.HexColor('#ffd6e8')))
    elements.append(Spacer(1, 0.2*cm))
    elements.append(make_table([
        ['Servicio',     comprobante['detalles_servicio']['servicio']],
        ['Especialista', empleado.nombre if empleado else 'Por asignar'],
        ['Fecha',        cita.fecha_hora_inicio.strftime('%d de %B de %Y')],
        ['Hora inicio',  cita.fecha_hora_inicio.strftime('%H:%M')],
        ['Hora fin',     cita.fecha_hora_fin.strftime('%H:%M')],
        ['Duración',     f'{servicio.duracion_minutos} min' if servicio else 'N/A'],
    ]))
    elements.append(Spacer(1, 0.5*cm))

    # ── Sección: Información de Pago ─────────────────────────────────────────
    elements.append(Paragraph('INFORMACIÓN DE PAGO', sec_s))
    elements.append(HRFlowable(width='100%', thickness=1, color=colors.HexColor('#ffd6e8')))
    elements.append(Spacer(1, 0.2*cm))
    pago_t = Table([
        ['Total del servicio', f'${float(cita.monto_total or 0):,.0f} COP'],
        ['Abono realizado',    comprobante['pago']['monto_abono']],
        ['Saldo pendiente',    f'${float(cita.saldo_pendiente or 0):,.0f} COP'],
    ], colWidths=[8*cm, 8*cm])
    pago_t.setStyle(TableStyle([
        ('FONTNAME',       (0,0), (0,-1), 'Helvetica-Bold'),
        ('FONTNAME',       (1,0), (1,-1), 'Helvetica-Bold'),
        ('FONTSIZE',       (0,0), (-1,-1), 10),
        ('TEXTCOLOR',      (0,0), (0,-1),  GREY),
        ('TEXTCOLOR',      (1,0), (1,-1),  BLACK),
        ('ROWBACKGROUNDS', (0,0), (-1,-1), [colors.white, LPINK]),
        ('TOPPADDING',     (0,0), (-1,-1), 8),
        ('BOTTOMPADDING',  (0,0), (-1,-1), 8),
        ('LEFTPADDING',    (0,0), (-1,-1), 10),
        ('ALIGN',          (1,0), (1,-1),  'RIGHT'),
        ('RIGHTPADDING',   (1,0), (1,-1),  10),
        ('BACKGROUND',     (0,2), (-1,2),  colors.HexColor('#fff0f6')),
        ('TEXTCOLOR',      (1,2), (1,2),   PINK),
        ('FONTSIZE',       (1,2), (1,2),   12),
    ]))
    elements.append(pago_t)
    elements.append(Spacer(1, 0.5*cm))

    # Nota y footer
    elements.append(Paragraph(
        '⚠️  Política de cancelación: debes cancelar con mínimo 2 horas de anticipación desde tu panel.',
        ParagraphStyle('nota', fontName='Helvetica', fontSize=9, textColor=GREY, alignment=TA_CENTER)
    ))
    elements.append(Spacer(1, 0.3*cm))
    elements.append(HRFlowable(width='100%', thickness=1, color=colors.HexColor('#ffd6e8')))
    elements.append(Spacer(1, 0.2*cm))
    elements.append(Paragraph(
        f'Generado el {datetime.now().strftime("%d/%m/%Y a las %H:%M")} · Rossmix Salón de Belleza',
        footer_s
    ))

    doc.build(elements)
    buffer.seek(0)

    filename = f"comprobante_{cita.codigo_reserva}_{cita.fecha_hora_inicio.strftime('%Y%m%d')}.pdf"
    logging.info(f"[PDF] Descarga iniciada: {filename}")
    return send_file(buffer, download_name=filename,
                     as_attachment=True, mimetype='application/pdf')


if __name__ == '__main__':
    app.run(debug=True)
